"""
api/services/outbound_report.py
=================================
日次 出庫レポート の データ構築 と Excel 生成。

VBA 「20260520_出庫レポート」 シート を Web 化:
  12 列構成:
    A 前月繰越     B 重量(=前月+月入荷)  C 当月出庫数(=B-D)  D 残高
    E 産地名       F 規格名 (=原料規格)   G 平均単価           H 当日出庫数
    I 商品規格名   J 出荷数 (raw kg)      K 小計 (= J SUM)     L 歩留まり

  ヘッダ:
    Row1: タイトル 「YYYY/M/D 出庫レポート」
    Row2: グループヘッダ
    Row3: サブヘッダ

  グループ化:
    1 (origin, raw_grade) = 主行 (A〜H 必ず 表示)
    今日 出庫あれば I〜L を 商品規格別 サブ行 で 展開:
      - 1 raw → N 商品 = N サブ行 (A〜H は 1 行目 だけ 値、 残り 空 = TypeA 結合相当)
      - 振替経由 (order_id IS NOT NULL): I=商品規格, J=raw qty, L=yield
      - 直接 (order_id IS NULL): I="(直接)", L=1.0

  今日 出庫が 無い (origin, raw_grade) でも 在庫 / 月内動き が あれば 表示
  (= "未動アイテム" / VBA でも 同じ)
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional

import psycopg
from psycopg.rows import dict_row

from api.services.grade_label import compact_grade_label


def _prev_period_str(d: date) -> str:
    """date(2026,5,20) → '2026-04'"""
    y, m = d.year, d.month
    if m == 1:
        return f"{y-1}-12"
    return f"{y}-{m-1:02d}"


def _month_str(d: date) -> str:
    """date(2026,5,20) → '2026-05'"""
    return f"{d.year}-{d.month:02d}"


def _month_start(d: date) -> date:
    return date(d.year, d.month, 1)


@dataclass
class ProductSubRow:
    """1 (origin, raw_grade, product_grade) の 当日 消化 行"""
    product_grade_label: str       # I 列 (例: "標準/A/L"、 直接の場合 "(直接)")
    # J 列 「商品出荷数」: 商品 換算 kg (= raw × yield)。 振替 で yield<1.0 の とき は
    # 出庫済 原料 と 異なる 値 に なる (= 真の 商品 出荷量)。 旧 名前 raw_qty_kg は
    # 誤称 で 互換性 の ため 維持 (2026-05-28 bug fix)。
    raw_qty_kg: Decimal            # J 列 (= 商品換算 kg、 product_qty_covered または raw if 直接)
    yield_applied: Decimal         # L 列 (1.0 = 直接、 <1.0 = 振替)


@dataclass
class ReportRow:
    """1 (origin, raw_grade) 行 のデータ"""
    origin_id: int
    origin_name: str               # E 列
    raw_grade_id: int
    raw_grade_label: str           # F 列
    prev_kg: Decimal = Decimal(0)  # A 列
    weight_kg: Decimal = Decimal(0)  # B 列 = prev + 当月 入荷
    month_out_kg: Decimal = Decimal(0)  # C 列 = B - D
    balance_kg: Decimal = Decimal(0)    # D 列 = B - 当月出庫
    avg_price: Decimal | None = None    # G 列
    today_out_kg: Decimal = Decimal(0)  # H 列
    sub_rows: list[ProductSubRow] = field(default_factory=list)  # I〜L


@dataclass
class ReportData:
    target_date: date
    crop_id: int
    crop_name: str
    rows: list[ReportRow] = field(default_factory=list)


# =============================================================================
# データ取得
# =============================================================================
async def build_report_data(
    conn: psycopg.AsyncConnection,
    target_date: date,
    crop_id: int,
) -> ReportData:
    """target_date 時点の 出庫レポート データ を 集計する。"""
    prev_period = _prev_period_str(target_date)
    month = _month_str(target_date)
    month_start = _month_start(target_date)

    # crop name
    async with conn.cursor(row_factory=dict_row) as cur:
        await cur.execute("SELECT name FROM crops WHERE id = %s", (crop_id,))
        crop_row = await cur.fetchone()
        crop_name = crop_row["name"] if crop_row else f"crop{crop_id}"

        # 全 lot + マスタ + 前月繰越 + 月内入荷 + 月内出庫 + 当日出庫
        await cur.execute("""
            SELECT
                il.id AS lot_id,
                il.total_kg,
                il.inbound_date,
                il.unit_price,
                p.origin_id,
                o.name AS origin_name,
                p.grade_id AS raw_grade_id,
                g.spec_type   AS raw_spec,
                g.grade_level AS raw_grade,
                g.size_label  AS raw_size,
                (SELECT counted_kg FROM stock_counts WHERE lot_id=il.id AND period=%(prev_period)s) AS prev_count,
                COALESCE((
                    SELECT SUM(quantity_kg) FROM outbound_records r
                    WHERE r.lot_id = il.id
                      AND r.outbound_date >= %(month_start)s
                      AND r.outbound_date <= %(target_date)s
                ), 0) AS month_out_kg
            FROM inbound_lots il
            JOIN products p ON p.id = il.product_id
            LEFT JOIN origins o ON o.id = p.origin_id
            JOIN grades g ON g.id = p.grade_id
            WHERE p.crop_id = %(crop_id)s
              AND il.archived_at IS NULL
        """, {
            "crop_id": crop_id, "prev_period": prev_period,
            "month_start": month_start, "target_date": target_date,
        })
        lots = await cur.fetchall()

        # 当日出庫 (lot 単位、 振替メタデータ含む)
        await cur.execute("""
            SELECT r.lot_id, r.quantity_kg, r.order_id, r.priority_used,
                   r.yield_applied, r.product_qty_covered,
                   oo.from_grade_id,
                   fg.spec_type   AS prod_spec,
                   fg.grade_level AS prod_grade,
                   fg.size_label  AS prod_size
            FROM outbound_records r
            JOIN inbound_lots il ON il.id = r.lot_id
            JOIN products p ON p.id = il.product_id
            LEFT JOIN outbound_orders oo ON oo.id = r.order_id
            LEFT JOIN grades fg ON fg.id = oo.from_grade_id
            WHERE p.crop_id = %s AND r.outbound_date = %s
        """, (crop_id, target_date))
        today_outs = await cur.fetchall()

    # group lots by (origin_id, raw_grade_id)
    groups: dict[tuple[int, int], dict] = defaultdict(lambda: {
        "lots": [],
        "prev_kg": Decimal(0),
        "month_in_kg": Decimal(0),
        "month_out_kg": Decimal(0),
        "weighted_price_sum": Decimal(0),
        "price_weight_sum": Decimal(0),
    })

    for lot in lots:
        key = (lot["origin_id"], lot["raw_grade_id"])
        g = groups[key]
        g["lots"].append(lot["lot_id"])

        if lot["prev_count"] is not None:
            g["prev_kg"] += Decimal(lot["prev_count"])
        if lot["inbound_date"].month == target_date.month and lot["inbound_date"].year == target_date.year \
                and lot["inbound_date"] <= target_date:
            g["month_in_kg"] += Decimal(lot["total_kg"])
        g["month_out_kg"] += Decimal(lot["month_out_kg"])
        # 平均単価 = lot.total_kg 加重平均
        if lot["unit_price"] is not None:
            g["weighted_price_sum"] += Decimal(lot["unit_price"]) * Decimal(lot["total_kg"])
            g["price_weight_sum"] += Decimal(lot["total_kg"])
        # group metadata (1 度だけ セット)
        if "origin_name" not in g:
            g["origin_name"] = lot["origin_name"] or ""
            g["raw_grade_label"] = compact_grade_label(lot["raw_spec"], lot["raw_grade"], lot["raw_size"])

    # 当日出庫 を group に 紐付け
    # 同 group 内 で 同じ product_grade + yield なら 1 サブ行 に 集約
    today_by_group: dict[tuple[int, int], dict[tuple[str, Decimal], Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal(0)))
    today_total_by_group: dict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal(0))

    # lot_id → (origin_id, raw_grade_id)
    lot_to_key: dict[int, tuple[int, int]] = {}
    for lot in lots:
        lot_to_key[lot["lot_id"]] = (lot["origin_id"], lot["raw_grade_id"])

    for t in today_outs:
        key = lot_to_key.get(t["lot_id"])
        if key is None:
            continue
        # H 列 「当日出庫数」 = 原料 (raw) 合計
        today_total_by_group[key] += Decimal(t["quantity_kg"])
        if t["order_id"] is not None and t["from_grade_id"] is not None:
            prod_label = compact_grade_label(t["prod_spec"], t["prod_grade"], t["prod_size"])
            yld = Decimal(t["yield_applied"] or 1)
        else:
            prod_label = "(直接)"
            yld = Decimal(1)
        sub_key = (prod_label, yld)
        # J 列 「商品出荷数」 = 商品 換算 (= raw × yield)。 振替 record は
        # product_qty_covered が セット 済、 直接出庫 は NULL なので quantity_kg
        # (= raw、 yield=1.0 で 商品=原料 と 等価) で 代用。 2026-05-28 bug fix:
        # 旧 logic は ここ で quantity_kg (raw) を 入れて いて 列 header と 不一致 だった。
        if t["product_qty_covered"] is not None:
            product_kg = Decimal(t["product_qty_covered"])
        else:
            product_kg = Decimal(t["quantity_kg"])  # 直接出庫 = yield 1.0
        today_by_group[key][sub_key] += product_kg

    # ReportRow を 組み立て
    rows: list[ReportRow] = []
    for key, g in groups.items():
        # 表示対象 判定: prev>0 or 月内in>0 or 月内out>0
        if g["prev_kg"] == 0 and g["month_in_kg"] == 0 and g["month_out_kg"] == 0:
            continue
        weight = g["prev_kg"] + g["month_in_kg"]
        balance = weight - g["month_out_kg"]
        avg_price = (
            g["weighted_price_sum"] / g["price_weight_sum"]
            if g["price_weight_sum"] > 0 else None
        )
        row = ReportRow(
            origin_id=key[0], origin_name=g.get("origin_name", ""),
            raw_grade_id=key[1], raw_grade_label=g.get("raw_grade_label", ""),
            prev_kg=g["prev_kg"], weight_kg=weight,
            month_out_kg=g["month_out_kg"], balance_kg=balance,
            avg_price=avg_price,
            today_out_kg=today_total_by_group.get(key, Decimal(0)),
        )
        # 当日 サブ行
        for (prod_label, yld), raw_qty in today_by_group.get(key, {}).items():
            row.sub_rows.append(ProductSubRow(
                product_grade_label=prod_label,
                raw_qty_kg=raw_qty,
                yield_applied=yld,
            ))
        # サブ行 を 商品規格 で ソート (見やすさ)
        row.sub_rows.sort(key=lambda s: s.product_grade_label)
        rows.append(row)

    # 並び順: 産地 → 規格
    rows.sort(key=lambda r: (r.origin_name, r.raw_grade_label))

    return ReportData(target_date=target_date, crop_id=crop_id, crop_name=crop_name, rows=rows)


# =============================================================================
# Excel 生成
# =============================================================================
def build_report_xlsx(data: ReportData) -> bytes:
    """ReportData から openpyxl で 12 列 Excel を 生成。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    d = data.target_date
    ws.title = f"{d.strftime('%Y%m%d')}_出庫レポート"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(border_style='thin', color='888888')
    thick = Side(border_style='medium', color='000000')
    br = Border(top=thin, bottom=thin, left=thin, right=thin)
    fill_hdr_a = PatternFill('solid', fgColor='E0F0FF')   # 在庫分析
    fill_hdr_b = PatternFill('solid', fgColor='FFE8D0')   # 当日出庫まとめ
    fill_hdr_c = PatternFill('solid', fgColor='E8F0D0')   # 商品規格別

    # Row 1: タイトル
    title = f"{d.year}/{d.month}/{d.day} 出庫レポート ({data.crop_name})"
    ws.cell(1, 1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    # Row 2: グループヘッダ
    group_hdr = [
        ("A", 4, "在庫分析",                       fill_hdr_a),
        ("E", 4, "当日出庫まとめ",                  fill_hdr_b),
        ("I", 4, "当日原料規格別商品出荷数",         fill_hdr_c),
    ]
    for col_letter, span, label, fill in group_hdr:
        from openpyxl.utils import column_index_from_string
        col = column_index_from_string(col_letter)
        c = ws.cell(2, col, value=label)
        c.font = bold; c.fill = fill
        c.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)

    # Row 3: サブヘッダ
    sub_hdr = [
        "前月繰越", "重量", "当月出庫数", "残高",
        "産地名", "規格名", "平均単価", "当日出庫数",
        "規格名", "出荷数", "小計", "歩どまり",
    ]
    for i, label in enumerate(sub_hdr, start=1):
        c = ws.cell(3, i, value=label)
        c.font = bold
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = br
        if i <= 4:  c.fill = fill_hdr_a
        elif i <= 8: c.fill = fill_hdr_b
        else:        c.fill = fill_hdr_c

    # データ行
    row_idx = 4
    for r in data.rows:
        # メイン行 + サブ行 (= 商品規格 別)
        sub_count = max(1, len(r.sub_rows))
        first_row = row_idx
        last_row = row_idx + sub_count - 1

        # A〜H は 最初の行 だけ 値、 残り 空 (= TypeA 縦結合 相当 = フロントで rowSpan)
        ws.cell(first_row, 1, value=float(r.prev_kg))
        ws.cell(first_row, 2, value=float(r.weight_kg))
        ws.cell(first_row, 3, value=float(r.month_out_kg))
        ws.cell(first_row, 4, value=float(r.balance_kg))
        ws.cell(first_row, 5, value=r.origin_name)
        ws.cell(first_row, 6, value=r.raw_grade_label)
        ws.cell(first_row, 7, value=float(r.avg_price) if r.avg_price is not None else None)
        ws.cell(first_row, 8, value=float(r.today_out_kg))

        # 結合 (sub_count > 1 のとき A〜H + K を 縦結合)
        if sub_count > 1:
            for col in [1, 2, 3, 4, 5, 6, 7, 8, 11]:
                ws.merge_cells(start_row=first_row, start_column=col, end_row=last_row, end_column=col)

        # 小計 (K) = sum of J for this group
        ws.cell(first_row, 11, value=f"=SUM(J{first_row}:J{last_row})")

        # I/J/L = サブ行
        if r.sub_rows:
            for s_idx, s in enumerate(r.sub_rows):
                rr = first_row + s_idx
                ws.cell(rr, 9, value=s.product_grade_label)
                ws.cell(rr, 10, value=float(s.raw_qty_kg))
                ws.cell(rr, 12, value=float(s.yield_applied))
        # else: I/J/L 空 (= 未動アイテム)

        # 罫線
        for rr in range(first_row, last_row + 1):
            for col in range(1, 13):
                ws.cell(rr, col).border = br
        # 数値書式
        for col in [1, 2, 3, 4, 8, 10, 11]:
            for rr in range(first_row, last_row + 1):
                cv = ws.cell(rr, col).value
                if cv is not None:
                    ws.cell(rr, col).number_format = '#,##0.0'
        # 単価
        if r.avg_price is not None:
            ws.cell(first_row, 7).number_format = '"¥"#,##0'
        # 歩留
        for rr in range(first_row, last_row + 1):
            v = ws.cell(rr, 12).value
            if v is not None:
                ws.cell(rr, 12).number_format = '0.0000'

        row_idx = last_row + 1

    # 列幅
    widths = [10, 10, 12, 10, 10, 16, 10, 12, 14, 10, 10, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 行高 (Row 3 サブヘッダ)
    ws.row_dimensions[3].height = 30

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
