"""
api/services/calendar_excel.py
================================
日次カレンダー 紙レポート の Excel (.xlsx) 出力。
CalendarPrintPage.tsx のレイアウトをサーバ側で 再現し、 openpyxl で xlsx 生成する。

レイアウト:
  - タイトル行 (作物 + 月)
  - ヘッダ行: 仕入先 / 産地 / 規格 / ケース / kg/CS / 数量 / 単価 / 入荷日 /
              消費税 / 合計金額 / 前払日 / 前払金額 / 後払日 / 後払金額 /
              前月繰越 / 当月入荷 / 当月出庫 / 当月在庫 / 在庫評価額 /
              日付列 (1..N)
  - データ行 (1 ロット 1 行)
  - 合計行
  - セルコメント (脚注 リスト + 該当セル に色付け)

戻り値: BytesIO (xlsx バイナリ)
"""
from __future__ import annotations

from decimal import Decimal
from datetime import date
from io import BytesIO
from typing import Any

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# プロジェクト ルート / ロゴ パス。 backend 起動時 の cwd に 依存 しない よう 絶対パス で 解決
_LOGO_PATH = Path(__file__).resolve().parent.parent.parent / "frontend" / "public" / "logo.png"


# 色 (CalendarPrintPage の COLOR と同等)
COLOR_HEADER_BLUE   = "9BC2E6"   # identity 列 ヘッダ
COLOR_HEADER_YELLOW = "FFFF00"   # tax/grand_total 列 ヘッダ
COLOR_HEADER_GREEN  = "A9D08E"   # summary 列 ヘッダ
COLOR_HEADER_DAY    = "FCE4D6"   # 日付列 ヘッダ + コメント付セル
COLOR_BODY_BLUE     = "DDEBF7"   # identity 列 body
COLOR_TEXT_DARK     = "1F4E79"
COLOR_RED           = "C0392B"


# ColDef: (id, label, group, numeric, width)
#   幅は openpyxl の文字単位 (1 = 約 7 px)。 日本語 1 文字 ≒ 2 char 換算で余裕を持たせる。
ALWAYS_COLS = [
    ("supplier",    "仕入先",    "identity", False, 22),  # 'みどり物産(株)' 等
    ("origin",      "産地",      "identity", False, 11),  # '和歌山県' 等
    ("spec",        "規格",      "identity", False, 13),  # '加工品 A 2L' 等
    ("cases",       "ケース",    "identity", True,   7),
    ("kg_per_case", "kg/CS",    "identity", True,   8),
    ("total_kg",    "数量(kg)", "identity", True,  10),
    ("unit_price",  "単価",      "identity", True,  11),
    ("inbound_date","入荷日",    "identity", False, 12),
]
OPTIONAL_COLS = [
    ("tax",            "消費税(8%)", "tax",     True,  12),
    ("grand_total",    "合計金額",   "tax",     True,  15),
    ("prepay_date",    "前払日",     "prepay",  False, 12),
    ("prepay_amount",  "前払金額(円)", "prepay",  True,  15),
    ("postpay_date",   "後払日",     "postpay", False, 12),
    ("postpay_amount", "後払金額(円)", "postpay", True,  15),
    ("carryover",      "前月繰越",   "summary", True,  11),
    ("inbound_kg",     "当月入荷",   "summary", True,  11),
    ("outbound",       "当月出庫",   "summary", True,  11),
    ("end_kg",         "当月在庫",   "summary", True,  11),
    ("end_value",      "在庫評価額", "summary", True,  15),
    # ─── 月末棚卸 (紙レポート と 同 仕様、 identity 色帯) ───
    ("stocktake_kg",   "棚卸数",     "stocktake", True,  10),
    ("stocktake_diff", "差数",       "stocktake", True,  10),
    ("stocktake_note", "差数原因",   "stocktake", False, 18),
]
ALL_COLS = ALWAYS_COLS + OPTIONAL_COLS

# 金額(¥)表示する列の id 集合
YEN_COLS = {"unit_price", "tax", "grand_total", "prepay_amount", "postpay_amount", "end_value"}

TAX_RATE = Decimal("0.08")


def _n(v: Any) -> Decimal:
    """Decimal 化 (None/空文字は 0)。"""
    if v is None or v == "":
        return Decimal(0)
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal(0)


def _format_spec(lot: dict, crop_id: int | None = None) -> str:
    """規格 表示 文字列 を 組み立て。
    通常: spec_type (≠'標準') + grade_level (≠'-') + size_label (≠'-') を 連結。
    長芋 (crop_id=3): spec_type 単独 (= 等級 'AB込み' を 規格 化 した 業務 ルール、
    migration 086 で データ 移行 済)。 spec_type が 空/'標準' の とき は 旧 ルール で
    fallback (= 加工品 等 で 連結 表示 を 維持)。
    """
    parts = []
    spec_type = lot.get("spec_type")
    grade = lot.get("grade_level")
    size = lot.get("size_label")
    if crop_id == 3 and spec_type and spec_type != "標準":
        # 長芋: spec_type のみ (= 'AB込み' / '加工品' / 'ニュースター(姫)' 等)
        return spec_type
    if spec_type and spec_type != "標準":
        parts.append(spec_type)
    if grade and grade != "-":
        parts.append(grade)
    if size and size != "-":
        parts.append(size)
    return "".join(parts) or "—"


# Body セル に 色塗りつぶし する 列 (user 指定 2026-05-25)。 残り の 列 は 白。
#   仕入先 / 入荷日 (identity の 一部)
#   消費税 / 合計金額 (tax)
#   棚卸数 / 差数 / 差数原因 (stocktake 3 列)
BODY_FILLED_COLS = {
    "supplier", "inbound_date",
    "tax", "grand_total",
    "stocktake_kg", "stocktake_diff", "stocktake_note",
}


def _insert_logo(ws, anchor_cell: str) -> None:
    """シート の タイトル行 に ロゴ画像 を 挿入。 ファイル 不在 でも 例外 を 出さない (best effort)。

    画像 は anchor_cell の 左上 に float 配置。 row 1 height = 60 想定 で、
    高さ を 50 px に 縮小 (アスペクト 220:105 維持 → 幅 約 105 px)。
    """
    if not _LOGO_PATH.exists():
        return
    try:
        img = XLImage(str(_LOGO_PATH))
        # アスペクト 維持 (220:105 ≈ 2.10) で 高さ 50 px → 幅 約 105 px
        img.height = 50
        img.width = int(50 * 220 / 105)  # ≈ 105
        ws.add_image(img, anchor_cell)
    except Exception:
        # 画像 描画 失敗 は xlsx 出力 全体 を 落とさない (任意装飾 扱い)
        pass


def _group_color(group: str, is_header: bool, cid: str | None = None) -> str | None:
    if is_header:
        return {
            "identity":  COLOR_HEADER_BLUE,
            "tax":       COLOR_HEADER_YELLOW,
            "prepay":    COLOR_HEADER_GREEN,
            "postpay":   COLOR_HEADER_GREEN,
            "summary":   COLOR_HEADER_GREEN,
            "stocktake": COLOR_HEADER_BLUE,   # 紙レポート で identity と同 色帯
        }.get(group)
    # body — BODY_FILLED_COLS に 含まれる cid のみ 着色
    if cid is None or cid not in BODY_FILLED_COLS:
        return None
    return {
        "identity":  COLOR_BODY_BLUE,
        "tax":       COLOR_HEADER_YELLOW,
        "stocktake": COLOR_BODY_BLUE,
    }.get(group)


def _cell_value(col_id: str, lot: dict, crop_id: int | None = None) -> Any:
    """各列の値計算。 紙レポート (CalendarPrintPage.cellContent) と同じロジック。
    crop_id: 規格 列 の 表示 を crop 別 ルール (長芋 = spec_type 単独) に 切り替える。
    """
    if col_id == "supplier":     return lot.get("supplier_name") or ""
    if col_id == "origin":       return lot.get("origin_name") or ""
    if col_id == "spec":         return _format_spec(lot, crop_id=crop_id)
    if col_id == "cases":
        kpc = _n(lot.get("kg_per_case"))
        if kpc <= 0: return None
        return float(_n(lot.get("total_kg")) / kpc)
    if col_id == "kg_per_case":  return float(_n(lot.get("kg_per_case"))) or None
    if col_id == "total_kg":     return float(_n(lot.get("total_kg")))
    if col_id == "unit_price":   return float(_n(lot.get("unit_price"))) or None
    if col_id == "inbound_date":
        d = lot.get("inbound_date")
        return d if d else None
    if col_id == "tax":
        sub = _n(lot.get("total_kg")) * _n(lot.get("unit_price"))
        return float(round(sub * TAX_RATE)) if sub > 0 else None
    if col_id == "grand_total":
        sub = _n(lot.get("total_kg")) * _n(lot.get("unit_price"))
        if sub <= 0: return None
        tx = sub * TAX_RATE
        br = _n(lot.get("brokerage_fee"))
        fr = _n(lot.get("freight_fee"))
        return float(round(sub + tx + br + fr))
    if col_id == "prepay_date":
        d = lot.get("prepay_date")
        return d if d else None
    if col_id == "prepay_amount":
        v = _n(lot.get("prepay_amount"))
        return float(v) if v > 0 else None
    if col_id == "postpay_date":
        d = lot.get("postpay_date")
        return d if d else None
    if col_id == "postpay_amount":
        v = _n(lot.get("postpay_amount"))
        return float(v) if v > 0 else None
    if col_id == "carryover":    return float(_n(lot.get("carryover_kg")))
    if col_id == "inbound_kg":
        v = _n(lot.get("inbound_kg"))
        return float(v) if v > 0 else None
    if col_id == "outbound":     return float(_n(lot.get("outbound_kg")))
    if col_id == "end_kg":       return float(_n(lot.get("end_kg")))
    if col_id == "end_value":
        up = _n(lot.get("unit_price"))
        ek = _n(lot.get("end_kg"))
        return float(round(up * ek)) if up > 0 else None
    # ─── 月末棚卸 ───
    if col_id == "stocktake_kg":
        v = lot.get("stocktake_kg")
        return float(_n(v)) if v is not None else None
    if col_id == "stocktake_diff":
        v = lot.get("stocktake_diff")
        if v is None: return None
        dv = float(_n(v))
        return dv if dv != 0 else None    # 差数 0 は 表示しない (紙と同じ)
    if col_id == "stocktake_note":
        return lot.get("stocktake_note") or None
    return None


CALENDAR_SHEET_NAME = "仕入管理台帳"
ENTRIES_RAW_SHEET_NAME = "棚卸エントリ_raw"


def build_calendar_xlsx(
    calendar_data: dict,
    title: str,
    days: int,
    inventory_entries_summary: list[dict] | None = None,
    inventory_entries_raw: list[dict] | None = None,
) -> bytes:
    """カレンダー API のレスポンス (dict) から xlsx バイナリを生成。

    calendar_data: GET /calendar の response.dict() 相当
      ・month, days_in_month, lots[], crop_name, prepared_at
    title: シート/ヘッダ用 タイトル (例: 'みどり物産事業2部5月仕入管理台帳「通常」')
    days: 表示する日数 (1..N、 通常 N=月末日)
    inventory_entries_summary: 在庫紐づきなし棚卸 を (大分類, 小分類, 産地, 規格) で
      集計した行リスト。 各要素 = {category_major, category_minor, origin, spec,
      cases, total_kg}。 None または 空配列 なら 棚卸セクションは生成するが「該当なし」表示。
    inventory_entries_raw: 集計表 を SUMIFS 数式 化 する 用 の raw 行 (= 上記 summary
      の GROUP BY 前)。 None なら 棚卸エントリ_raw シート 自体 を 出さ ず、 商品/半製品
      集計 は 直値 fallback (2026-05-30 追加)。
    """
    # crop_id を 取得 (= 表示/集計 ルール の 切替 用、 長芋 で 規格 列 単独 表示)。
    # calendar_data には CalendarView.crop_id が 含まれる (= main.py /calendar/export.xlsx
    # で .model_dump(mode="json") して 渡されて くる)。
    crop_id = calendar_data.get("crop_id")
    wb = Workbook()
    ws = wb.active
    ws.title = CALENDAR_SHEET_NAME
    ws.sheet_view.showGridLines = False

    # 罫線 / 共通スタイル
    thin = Side(border_style="thin", color="999999")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    font_title  = Font(name="Yu Mincho", size=14, bold=True, color=COLOR_TEXT_DARK)
    font_header = Font(name="Yu Gothic", size=10, bold=True, color="3F3F00")
    font_body   = Font(name="Yu Gothic", size=10)
    font_total  = Font(name="Yu Gothic", size=10, bold=True)

    n_attr_cols = len(ALL_COLS)
    n_day_cols = days
    n_total_cols = n_attr_cols + n_day_cols
    last_col_letter = get_column_letter(n_total_cols)

    # ─── 1) タイトル行 (ロゴ + テキスト) ───
    # 行 1 は ロゴ画像 + 中央寄せ タイトル文字。 ロゴは A1 左寄せ で float、
    # タイトル文字は merge 範囲全体で 中央寄せ (ロゴ と 重ならない 位置 になる
    # ように merge 範囲 を A:end_col で 取り、 indent で 文字を 右へ オフセット)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = font_title
    # 左に ロゴ ぶん 余裕を 持たせる ため indent。 中央寄せ で 全体 として 見栄え 良く
    title_cell.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_total_cols)

    _insert_logo(ws, "A1")

    # ─── 2) ヘッダ行 (row=2) ───
    HEADER_ROW = 2
    for ci, (cid, label, group, numeric, width) in enumerate(ALL_COLS, start=1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=label)
        c.font = font_header
        c.alignment = center
        c.border = border_all
        color = _group_color(group, is_header=True)
        if color:
            c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.column_dimensions[get_column_letter(ci)].width = width

    # 日付列
    for d in range(1, n_day_cols + 1):
        ci = n_attr_cols + d
        c = ws.cell(row=HEADER_ROW, column=ci, value=d)
        c.font = font_header
        c.alignment = center
        c.border = border_all
        c.fill = PatternFill(start_color=COLOR_HEADER_DAY, end_color=COLOR_HEADER_DAY, fill_type="solid")
        ws.column_dimensions[get_column_letter(ci)].width = 5.5

    # ─── 3) データ行 ───
    BODY_START = HEADER_ROW + 1
    lots = calendar_data.get("lots", [])
    for li, lot in enumerate(lots):
        row = BODY_START + li
        # 属性列
        for ci, (cid, label, group, numeric, width) in enumerate(ALL_COLS, start=1):
            v = _cell_value(cid, lot, crop_id=crop_id)
            c = ws.cell(row=row, column=ci, value=v)
            c.font = font_body
            c.border = border_all
            if numeric:
                c.alignment = right
                if cid in YEN_COLS:
                    c.number_format = '"¥"#,##0;"¥"-#,##0'
                else:
                    c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
            elif cid == "inbound_date" or cid in ("prepay_date", "postpay_date"):
                c.alignment = center
                c.number_format = "yyyy-mm-dd"
            else:
                c.alignment = left
            color = _group_color(group, is_header=False, cid=cid)
            if color:
                c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # 日付列 (出庫量 + コメント色付け)
        daily = lot.get("daily") or {}
        comments = lot.get("comments") or {}
        for d in range(1, n_day_cols + 1):
            ci = n_attr_cols + d
            qty = _n(daily.get(str(d)))
            cmt = comments.get(str(d))
            v = float(round(qty)) if qty > 0 else None
            c = ws.cell(row=row, column=ci, value=v)
            c.font = font_body
            c.border = border_all
            c.alignment = center
            c.number_format = '#,##0'
            if cmt:
                # コメント付セルは 日付ヘッダ色 (#FCE4D6) で塗る + コメント (mouse hover で 表示)
                c.fill = PatternFill(start_color=COLOR_HEADER_DAY, end_color=COLOR_HEADER_DAY, fill_type="solid")
                from openpyxl.comments import Comment as XLComment
                c.comment = XLComment(cmt, "棚卸")

    # ─── 4) 合計行 (数式) ───
    #   各列の合計は SUM 式を入れて、 ユーザがデータセルを編集すると 自動再計算されるようにする。
    #   1 ロットも無い場合 (= データ行 0 件) は SUM(範囲) が空セルになるので 警告セル に切替え。
    total_row = BODY_START + len(lots)
    body_first = BODY_START
    body_last  = total_row - 1     # 最後のデータ行

    # 「合計対象でない」 列 (テキスト系 / 日付系) は除外、 SUM 式を入れない列は None
    SUM_COLS = {
        "cases", "total_kg", "tax", "grand_total",
        "prepay_amount", "postpay_amount",
        "carryover", "inbound_kg", "outbound", "end_kg", "end_value",
        "stocktake_kg",   # 棚卸数 のみ 合計 (差数/差数原因 は 合計しない、 紙と同じ)
    }
    for ci, (cid, label, group, numeric, width) in enumerate(ALL_COLS, start=1):
        col_letter = get_column_letter(ci)
        if cid == "supplier":
            c = ws.cell(row=total_row, column=ci, value="合計")
            c.alignment = center
        elif cid in SUM_COLS and len(lots) > 0:
            formula = f"=SUM({col_letter}{body_first}:{col_letter}{body_last})"
            c = ws.cell(row=total_row, column=ci, value=formula)
            c.alignment = right
            if cid in YEN_COLS:
                c.number_format = '"¥"#,##0;"¥"-#,##0'
            else:
                c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
        else:
            # 入荷日 / 前払日 / 後払日 / 単価 (これらは合計しない) → 空セル
            c = ws.cell(row=total_row, column=ci, value=None)
        c.font = font_total
        c.border = border_all
        c.fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")

    # 日付列 合計 (SUM 式)
    for d in range(1, n_day_cols + 1):
        ci = n_attr_cols + d
        col_letter = get_column_letter(ci)
        if len(lots) > 0:
            formula = f"=SUM({col_letter}{body_first}:{col_letter}{body_last})"
            c = ws.cell(row=total_row, column=ci, value=formula)
        else:
            c = ws.cell(row=total_row, column=ci, value=None)
        c.font = font_total
        c.alignment = center
        c.border = border_all
        c.fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")
        c.number_format = '#,##0'

    # ─── 5) コメント脚注 ───
    footnote_row = total_row + 2
    has_footnote = False
    for lot in lots:
        if lot.get("comments"):
            has_footnote = True
            break
    if has_footnote:
        ws.cell(row=footnote_row, column=1, value="※ コメント (色付セル に対応)").font = Font(name="Yu Gothic", size=10, bold=True, color=COLOR_TEXT_DARK)
        ws.merge_cells(start_row=footnote_row, start_column=1, end_row=footnote_row, end_column=n_total_cols)
        footnote_row += 1
        idx = 0
        month_str = calendar_data.get("month", "")
        mm = month_str[5:7] if month_str else ""
        for lot in lots:
            cmts = lot.get("comments") or {}
            if not cmts: continue
            for d_str in sorted(cmts.keys(), key=int):
                idx += 1
                text = f"{idx}. (整理番号 {lot.get('lot_code') or ('#'+str(lot.get('lot_id')))} / {mm}月{d_str}日) {cmts[d_str]}"
                ws.cell(row=footnote_row, column=1, value=text).font = Font(name="Yu Gothic", size=9)
                ws.merge_cells(start_row=footnote_row, start_column=1, end_row=footnote_row, end_column=n_total_cols)
                footnote_row += 1

    # ─── 印刷設定 ───
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    # ヘッダ行を 各ページの上部に繰り返し
    ws.print_title_rows = "1:2"
    # 1 列目を 各ページの左に繰り返し
    ws.print_title_cols = "A:A"
    # 印刷範囲
    ws.print_area = f"A1:{last_col_letter}{footnote_row}"
    # 行高
    ws.row_dimensions[1].height = 60   # タイトル + ロゴ 用 に 高め
    ws.row_dimensions[HEADER_ROW].height = 32
    # 凍結 は しない (user 指定 2026-05-25)

    # ─── 2 枚目: 在庫集計 + 商品・半製品等集計 シート ───
    _build_summary_sheet(
        wb,
        lots=lots,
        inventory_entries_summary=inventory_entries_summary or [],
        inventory_entries_raw=inventory_entries_raw or [],
        title=title,
        month=calendar_data.get("month", ""),
        crop_id=crop_id,
    )

    # ─── 3 枚目 (hidden): 棚卸エントリ raw — 商品/半製品 集計 SUMIFS 参照 元 ───
    if inventory_entries_raw:
        _build_entries_raw_sheet(wb, inventory_entries_raw)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =============================================================================
# 2 枚目シート: 産地・規格別 集計 (在庫一覧 + 商品・半製品等集計)
# =============================================================================
# レイアウト (オフィス向け 上品スタイル):
#   行 1   タイトル (淡いダークブルー、 ミンチョウ)
#   行 3   セクション見出し 「在庫一覧 (月末時点)」
#   行 4   表ヘッダ: 産地 / 規格 / ケース数 / 総重量(kg) / 金額(¥) / 平均単価(¥/kg)
#   行 5..  データ行 (zebra: 偶数行 薄グレー)
#   合計行
#   空行 x2
#   セクション見出し 「商品・半製品等集計」 (大分類/小分類 で 小表 分割)
#   各小表: 帯見出し ▼ 大分類 / 小分類 → ヘッダ → データ → 小計
#   全体合計 (金額/平均単価 列は 手動入力 想定 で 空欄)

SUMMARY_TITLE_FILL = "1F4E79"    # 濃紺
SUMMARY_SECTION_FILL = "D9E1F2"  # 淡いブルー (見出し帯)
SUMMARY_HEADER_FILL = "9BC2E6"   # 既存と統一
SUMMARY_ZEBRA_FILL = "F2F7FC"    # 偶数行
SUMMARY_TOTAL_FILL = "FFFBE6"    # 合計行 (既存と統一)
SUMMARY_BORDER_GRAY = "B0B7BE"


def _spec_combined(spec: str | None, sub_spec: str | None) -> str:
    """規約: sub_spec_text が あれば spec_text を 置換、 無ければ そのまま (集計表 系)。"""
    s = (sub_spec or "").strip()
    if s:
        return s
    return (spec or "").strip() or "—"


def _build_summary_sheet(
    wb: Workbook,
    lots: list[dict],
    inventory_entries_summary: list[dict],
    inventory_entries_raw: list[dict],
    title: str,
    month: str,
    crop_id: int | None = None,
) -> None:
    ws = wb.create_sheet("在庫集計")
    ws.sheet_view.showGridLines = False

    thin = Side(border_style="thin", color=SUMMARY_BORDER_GRAY)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    font_title   = Font(name="Yu Mincho", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Yu Gothic", size=12, bold=True, color=COLOR_TEXT_DARK)
    font_header  = Font(name="Yu Gothic", size=10, bold=True, color="1F2A44")
    font_body    = Font(name="Yu Gothic", size=10)
    font_total   = Font(name="Yu Gothic", size=10, bold=True, color="1F2A44")
    font_note    = Font(name="Yu Gothic", size=9, italic=True, color="6B7280")

    # 表 (在庫一覧 / 商品・半製品等集計) は 共に 6 列。 タイトル と セクション 帯 を
    # 6 列 で 揃える (= 旧 8 列設定 で 右端 2 列 が body 無し で 色塗り だけ 出る 違和感
    # を 解消)。
    TOTAL_COLS = 6

    # ─── タイトル ───
    mm = month[5:7] if month else ""
    yyyy = month[:4] if month else ""
    title_text = f"在庫・棚卸 集計表  ({yyyy}年{int(mm) if mm else ''}月)" if mm else "在庫・棚卸 集計表"
    ws.cell(row=1, column=1, value=title_text).font = font_title
    # ロゴ ぶん 左に 余白 を 取って 中央寄せ (indent 6 ≒ ロゴ幅 を 避ける)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center", indent=6)
    ws.cell(row=1, column=1).fill = PatternFill(
        start_color=SUMMARY_TITLE_FILL, end_color=SUMMARY_TITLE_FILL, fill_type="solid",
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    ws.row_dimensions[1].height = 60  # ロゴ + タイトル 用 に 高め

    _insert_logo(ws, "A1")

    # 元タイトル (台帳名) を 小さく サブタイトル として
    ws.cell(row=2, column=1, value=title).font = font_note
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    ws.row_dimensions[2].height = 16

    cur_row = 4  # 空行 1 つ あけて 在庫セクション

    # ═══ セクション 1: 在庫一覧 ═══
    cur_row = _render_section_header(
        ws, cur_row, TOTAL_COLS,
        text="■ 在庫一覧 (月末時点 / 産地・規格 別 集計)",
        font=font_section,
    )
    stock_columns = [
        ("origin",      "産地",         16, False),
        ("spec",        "規格",         20, False),
        ("cases",       "ケース数",     12, True),
        ("total_kg",    "総重量 (kg)",  14, True),
        ("amount",      "金額 (¥)",     16, True),
        ("unit_price",  "平均単価 (¥/kg)", 16, True),
    ]
    # 表ヘッダ
    for ci, (cid, label, width, _numeric) in enumerate(stock_columns, start=1):
        c = ws.cell(row=cur_row, column=ci, value=label)
        c.font = font_header
        c.alignment = center
        c.border = border_all
        c.fill = PatternFill(start_color=SUMMARY_HEADER_FILL, end_color=SUMMARY_HEADER_FILL, fill_type="solid")
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[cur_row].height = 24
    cur_row += 1

    # 在庫集計: (origin, spec) で groupby、 end_kg > 0 のロットのみ対象。
    # 2026-05-30 修正: 数式 が 参照 す べき は **当月在庫 (= end_kg、 列 R)**。
    # 旧 私 の 実装 は 入荷時 total_kg (列 F) を 参照 → 出庫 後 は ずれて 「総重量
    # 合計 ≠ 当月在庫 合計」 に なる bug だった。
    #   ・総重量      = SUM(end_kg)            (R 列)
    #   ・ケース       = SUM(end_kg / kpc)     (R / E、 SUMPRODUCT + IFERROR で 0 除算 防止)
    #   ・金額        = SUM(end_kg × unit_price)  (R × G)
    # 仕入管理台帳 列 index: A=supplier B=origin C=spec D=cases(入荷時) E=kpc F=total_kg(入荷時)
    #                       G=unit_price ... R=end_kg (= 当月在庫)
    stock_rows = _aggregate_stock_by_origin_spec(lots, crop_id=crop_id)
    cal_first = 3   # HEADER_ROW=2 → BODY_START=3
    cal_last  = 3 + max(0, len(lots) - 1)
    cal_sheet = f"'{CALENDAR_SHEET_NAME}'"
    cal_origin_r = f"{cal_sheet}!$B${cal_first}:$B${cal_last}"
    cal_spec_r   = f"{cal_sheet}!$C${cal_first}:$C${cal_last}"
    cal_kpc_r    = f"{cal_sheet}!$E${cal_first}:$E${cal_last}"
    cal_price_r  = f"{cal_sheet}!$G${cal_first}:$G${cal_last}"
    cal_end_r    = f"{cal_sheet}!$R${cal_first}:$R${cal_last}"   # 当月在庫
    body_first = cur_row
    if stock_rows:
        for i, r in enumerate(stock_rows):
            zebra = (i % 2 == 1)
            origin_ref = f"$A{cur_row}"
            spec_ref   = f"$B{cur_row}"
            # ケース: end_kg / kpc を 個別 計算 して 合計。
            # 注: 旧 IFERROR(R/E,0) パターン だと Excel が _xlfn.SINGLE で 自動 ラップ し
            # 配列 1 要素 だけ 評価 → 結果 0 に なる bug (= 2026-05-31 user 報告)。
            # 「(E>0) で マスク + 0除算 防止 で E+(E=0) で 分母 を 1 に 置換」 で 回避。
            # E=0 行 は (E>0)=0 で マスクされて contribution 0、 安全。
            cases_formula  = (f"=SUMPRODUCT(({cal_kpc_r}>0)*{cal_end_r}/"
                              f"({cal_kpc_r}+({cal_kpc_r}=0))*"
                              f"({cal_origin_r}={origin_ref})*({cal_spec_r}={spec_ref}))"
                              ) if len(lots) > 0 else 0
            # 総重量: SUM(end_kg) — SUMIFS で OK
            kg_formula     = (f"=SUMIFS({cal_end_r},{cal_origin_r},{origin_ref},"
                              f"{cal_spec_r},{spec_ref})") if len(lots) > 0 else 0
            # 金額: SUM(end_kg × unit_price)
            amount_formula = (f"=SUMPRODUCT({cal_end_r}*{cal_price_r}*"
                              f"({cal_origin_r}={origin_ref})*({cal_spec_r}={spec_ref}))"
                              ) if len(lots) > 0 else 0
            unit_price_formula = f"=IF(D{cur_row}>0,E{cur_row}/D{cur_row},\"\")"
            values = [
                r["origin"], r["spec"],
                cases_formula, kg_formula, amount_formula,
                unit_price_formula,
            ]
            for ci, (cid, _label, _w, numeric) in enumerate(stock_columns, start=1):
                v = values[ci - 1]
                c = ws.cell(row=cur_row, column=ci, value=v)
                c.font = font_body
                c.border = border_all
                if numeric:
                    c.alignment = right
                    if cid in ("amount", "unit_price"):
                        c.number_format = '"¥"#,##0'
                    elif cid == "cases":
                        c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
                    else:
                        c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
                else:
                    c.alignment = left
                if zebra:
                    c.fill = PatternFill(start_color=SUMMARY_ZEBRA_FILL, end_color=SUMMARY_ZEBRA_FILL, fill_type="solid")
            cur_row += 1
    else:
        c = ws.cell(row=cur_row, column=1, value="(該当する在庫はありません)")
        c.font = font_note
        c.alignment = center
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(stock_columns))
        cur_row += 1

    # 合計行
    body_last = cur_row - 1
    _render_total_row(
        ws, cur_row, stock_columns,
        body_first=body_first, body_last=body_last,
        has_body=bool(stock_rows),
        label_col=1, label_text="合計",
        font_total=font_total, border_all=border_all,
        weighted_avg_cols={"unit_price": ("amount", "total_kg")},
    )
    cur_row += 1

    # ═══ セクション 2: 商品・半製品等 集計 ═══
    # (大分類, 小分類) ごとに 小表に分割。 各小表に 見出し帯 + ヘッダ + データ + 小計。
    # ※ 旧名「棚卸集計 (在庫に紐づかない / 最新エントリ)」 + 注記 を ユーザー 視点
    #   (= システム側 概念 を 出さない) で 簡素化 (2026-05-26)。
    cur_row += 2  # 空行 2 つ
    cur_row = _render_section_header(
        ws, cur_row, TOTAL_COLS,
        text="■ 商品・半製品等集計",
        font=font_section,
    )
    cur_row += 1  # セクション帯 直後 に 1 行 余白

    # 小表用 列定義 (在庫セクションと別。 大分類/小分類 は 帯見出し に出すので
    # 表ヘッダには含めない)
    entry_columns = [
        ("origin",     "産地",           16, False),
        ("spec",       "規格",           20, False),
        ("cases",      "ケース数",       12, True),
        ("total_kg",   "総重量 (kg)",    14, True),
        ("amount",     "金額 (¥)",       16, True),
        ("unit_price", "平均単価 (¥/kg)", 16, True),
    ]
    # 列幅: 在庫セクションと共通 (列 1-6 は 既に設定済み)。 念のため 不足分のみ補填
    for ci, (_cid, _label, width, _numeric) in enumerate(entry_columns, start=1):
        col_letter = get_column_letter(ci)
        if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < width:
            ws.column_dimensions[col_letter].width = width

    # 小表用 帯見出し色 (在庫セクションの帯より 少し 違う 色味で 区別)
    SUBGROUP_FILL = "E2EFDA"  # 淡い緑帯
    font_subgroup = Font(name="Yu Gothic", size=11, bold=True, color=COLOR_TEXT_DARK)

    # (大分類, 小分類) で groupby
    grouped = _group_inventory_entries(inventory_entries_summary)

    # 各サブ表の小計行番号 (全体合計式 で 参照する)
    entry_subtotal_rows: list[int] = []

    if not grouped:
        c = ws.cell(row=cur_row, column=1, value="(該当データは ありません)")
        c.font = font_note
        c.alignment = center
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=TOTAL_COLS)
        cur_row += 1
    else:
        for group_idx, ((cmaj, cmin), rows) in enumerate(grouped):
            # 帯見出し: 「▼ 大分類 / 小分類」
            cmaj_label = cmaj or "(未分類)"
            cmin_label = cmin or "(未分類)"
            label_text = f"▼ {cmaj_label} / {cmin_label}    ({len(rows)} 件)"
            c = ws.cell(row=cur_row, column=1, value=label_text)
            c.font = font_subgroup
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.fill = PatternFill(start_color=SUBGROUP_FILL, end_color=SUBGROUP_FILL, fill_type="solid")
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(entry_columns))
            ws.row_dimensions[cur_row].height = 20
            cur_row += 1

            # 表ヘッダ
            for ci, (_cid, label, _w, _numeric) in enumerate(entry_columns, start=1):
                c = ws.cell(row=cur_row, column=ci, value=label)
                c.font = font_header
                c.alignment = center
                c.border = border_all
                c.fill = PatternFill(start_color=SUMMARY_HEADER_FILL, end_color=SUMMARY_HEADER_FILL, fill_type="solid")
            ws.row_dimensions[cur_row].height = 22
            cur_row += 1

            # データ行。 金額 は 手動入力、 平均単価 は 「= 金額 ÷ 総重量」 数式。
            # 2026-05-30: cases / total_kg を SUMIFS で 棚卸エントリ_raw シート 参照 に。
            # raw シート が ある とき (= entries あり) のみ 数式、 ない なら 直値 fallback。
            body_first = cur_row
            # raw 参照 範囲 (= ヘッダ1行 + entries n 行)
            n_raw = len(inventory_entries_raw)
            raw_last = 1 + n_raw if n_raw > 0 else 1
            raw_sheet = f"'{ENTRIES_RAW_SHEET_NAME}'"
            raw_cmaj_r   = f"{raw_sheet}!$A$2:$A${raw_last}"
            raw_cmin_r   = f"{raw_sheet}!$B$2:$B${raw_last}"
            raw_origin_r = f"{raw_sheet}!$C$2:$C${raw_last}"
            raw_spec_r   = f"{raw_sheet}!$D$2:$D${raw_last}"
            raw_cases_r  = f"{raw_sheet}!$E$2:$E${raw_last}"
            raw_kg_r     = f"{raw_sheet}!$F$2:$F${raw_last}"
            # cmaj / cmin の string literal (None → 空文字)。 二重引用符 escape は SUMIFS 仕様
            cmaj_lit = (cmaj or "").replace('"', '""')
            cmin_lit = (cmin or "").replace('"', '""')
            for i, r in enumerate(rows):
                zebra = (i % 2 == 1)
                spec_display = _spec_combined(r.get("spec"), r.get("sub_spec"))
                cases_v = r.get("cases")
                total_kg_v = r.get("total_kg")
                unit_price_formula = f"=IF(D{cur_row}>0,E{cur_row}/D{cur_row},\"\")"
                origin_ref = f"$A{cur_row}"
                spec_ref   = f"$B{cur_row}"
                if n_raw > 0:
                    cases_cell = (f'=SUMIFS({raw_cases_r},{raw_cmaj_r},"{cmaj_lit}",'
                                  f'{raw_cmin_r},"{cmin_lit}",{raw_origin_r},{origin_ref},'
                                  f'{raw_spec_r},{spec_ref})')
                    kg_cell    = (f'=SUMIFS({raw_kg_r},{raw_cmaj_r},"{cmaj_lit}",'
                                  f'{raw_cmin_r},"{cmin_lit}",{raw_origin_r},{origin_ref},'
                                  f'{raw_spec_r},{spec_ref})')
                else:
                    cases_cell = float(cases_v) if cases_v is not None else None
                    kg_cell    = float(total_kg_v) if total_kg_v is not None else None
                values = [
                    r.get("origin") or "—",
                    spec_display,
                    cases_cell,
                    kg_cell,
                    None,  # 金額 (手動入力)
                    unit_price_formula,  # 平均単価 = 金額/総重量 (金額 入力後 自動算出)
                ]
                for ci, (cid, _label, _w, numeric) in enumerate(entry_columns, start=1):
                    v = values[ci - 1]
                    c = ws.cell(row=cur_row, column=ci, value=v)
                    c.font = font_body
                    c.border = border_all
                    if numeric:
                        c.alignment = right
                        if cid in ("amount", "unit_price"):
                            c.number_format = '"¥"#,##0'
                        elif cid == "cases":
                            c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
                        else:
                            c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
                    else:
                        c.alignment = left
                    if zebra:
                        c.fill = PatternFill(start_color=SUMMARY_ZEBRA_FILL, end_color=SUMMARY_ZEBRA_FILL, fill_type="solid")
                cur_row += 1

            # 小計行 (金額は SUM = 手動入力した値を 自動集計、 平均単価は 加重平均式)
            #   金額セルが 全部空 でも SUM 式 は 空 (= 0) を返すだけ で 害は ない
            body_last = cur_row - 1
            subtotal_row = cur_row
            _render_total_row(
                ws, cur_row, entry_columns,
                body_first=body_first, body_last=body_last,
                has_body=True,
                label_col=1, label_text="小計",
                font_total=font_total, border_all=border_all,
                weighted_avg_cols={"unit_price": ("amount", "total_kg")},
            )
            entry_subtotal_rows.append(subtotal_row)
            cur_row += 1

            # 小表間 空行 (最後のグループの後は付けない)
            if group_idx < len(grouped) - 1:
                cur_row += 1

    # ─── 棚卸 全体合計 (2 つ以上 サブ表 がある 場合のみ) ───
    if len(entry_subtotal_rows) >= 2:
        cur_row += 1  # 全体合計 直前 に 空行
        _render_grand_total_row(
            ws, cur_row, entry_columns,
            subtotal_rows=entry_subtotal_rows,
            label_col=1, label_text="全体合計",
            font_total=font_total, border_all=border_all,
            weighted_avg_cols={"unit_price": ("amount", "total_kg")},
        )
        cur_row += 1

    # ─── ページ設定 ───
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    # 凍結 は しない (user 指定 2026-05-25)


def _build_entries_raw_sheet(wb: Workbook, raw: list[dict]) -> None:
    """棚卸エントリ raw 行 を hidden シート に 書く。 商品/半製品 集計 の SUMIFS
    参照 元。 列: A=大分類 B=小分類 C=産地 D=規格(統合) E=ケース数 F=総重量
              G=ケース重量 H=名前 I=棚卸日。 全 行 ヘッダ込み。
    シート は state='hidden' で 配信 (= user が 必要時 unhide 可)。"""
    ws = wb.create_sheet(ENTRIES_RAW_SHEET_NAME)
    ws.sheet_state = "hidden"
    headers = ["大分類", "小分類", "産地", "規格", "ケース数", "総重量",
               "ケース重量", "名前", "棚卸日"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h).font = Font(name="Yu Gothic", size=10, bold=True)
    for ri, r in enumerate(raw, start=2):
        ws.cell(row=ri, column=1, value=r.get("category_major") or "")
        ws.cell(row=ri, column=2, value=r.get("category_minor") or "")
        ws.cell(row=ri, column=3, value=r.get("origin") or "")
        ws.cell(row=ri, column=4, value=_spec_combined(r.get("spec"), r.get("sub_spec")))
        ws.cell(row=ri, column=5, value=r.get("cases"))
        ws.cell(row=ri, column=6, value=r.get("total_kg"))
        ws.cell(row=ri, column=7, value=r.get("kg_per_case"))
        ws.cell(row=ri, column=8, value=r.get("name"))
        ws.cell(row=ri, column=9, value=r.get("inventory_date"))


def _render_section_header(ws, row: int, total_cols: int, text: str, font: Font) -> int:
    """セクション見出し帯 を 1 行 出力。 戻り値 = 次に書く 行番号。"""
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill(start_color=SUMMARY_SECTION_FILL, end_color=SUMMARY_SECTION_FILL, fill_type="solid")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    ws.row_dimensions[row].height = 22
    return row + 1


def _render_grand_total_row(
    ws, row: int, columns: list[tuple],
    subtotal_rows: list[int],
    label_col: int, label_text: str,
    font_total: Font, border_all: Border,
    weighted_avg_cols: dict[str, tuple[str, str]] | None = None,
) -> None:
    """連続して いない 複数の 小計行 を 参照 して 全体合計 を 出す。
    範囲 SUM では なく 個別セル を + で 連結 する (空行/見出しを 跨ぐ ため)。
    """
    weighted_avg_cols = weighted_avg_cols or {}
    cid_to_col_idx = {cid: i + 1 for i, (cid, *_rest) in enumerate(columns)}

    # 全体合計 は サブ表 と 区別 する ため 少し 濃い 色
    grand_fill = "FFE7A8"  # 山吹

    for ci, (cid, _label, _w, numeric) in enumerate(columns, start=1):
        col_letter = get_column_letter(ci)
        if ci == label_col:
            c = ws.cell(row=row, column=ci, value=label_text)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif numeric:
            if cid in weighted_avg_cols:
                num_cid, den_cid = weighted_avg_cols[cid]
                num_letter = get_column_letter(cid_to_col_idx[num_cid])
                den_letter = get_column_letter(cid_to_col_idx[den_cid])
                num_expr = "+".join(f"{num_letter}{r}" for r in subtotal_rows)
                den_expr = "+".join(f"{den_letter}{r}" for r in subtotal_rows)
                formula = f"=IF(({den_expr})>0,({num_expr})/({den_expr}),\"\")"
                c = ws.cell(row=row, column=ci, value=formula)
            else:
                formula = "=" + "+".join(f"{col_letter}{r}" for r in subtotal_rows)
                c = ws.cell(row=row, column=ci, value=formula)
            c.alignment = Alignment(horizontal="right", vertical="center")
            if cid in ("amount", "unit_price"):
                c.number_format = '"¥"#,##0'
            elif cid == "cases":
                c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
            else:
                c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
        else:
            c = ws.cell(row=row, column=ci, value=None)
        c.font = font_total
        c.border = border_all
        c.fill = PatternFill(start_color=grand_fill, end_color=grand_fill, fill_type="solid")
    ws.row_dimensions[row].height = 24


def _render_total_row(
    ws, row: int, columns: list[tuple],
    body_first: int, body_last: int, has_body: bool,
    label_col: int, label_text: str,
    font_total: Font, border_all: Border,
    weighted_avg_cols: dict[str, tuple[str, str]] | None = None,
    skip_sum_cols: set[str] | None = None,
) -> None:
    """合計行を 1 行出力。
    columns: [(cid, label, width, numeric), ...]
    weighted_avg_cols: { cid: (numerator_cid, denominator_cid) } — 重み付き平均
    skip_sum_cols: SUM 式を入れない 数値列 (空セル)
    """
    weighted_avg_cols = weighted_avg_cols or {}
    skip_sum_cols = skip_sum_cols or set()
    cid_to_col_idx = {cid: i + 1 for i, (cid, *_rest) in enumerate(columns)}

    for ci, (cid, _label, _w, numeric) in enumerate(columns, start=1):
        col_letter = get_column_letter(ci)
        if ci == label_col:
            c = ws.cell(row=row, column=ci, value=label_text)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif numeric and has_body and cid not in skip_sum_cols:
            if cid in weighted_avg_cols:
                num_cid, den_cid = weighted_avg_cols[cid]
                num_letter = get_column_letter(cid_to_col_idx[num_cid])
                den_letter = get_column_letter(cid_to_col_idx[den_cid])
                formula = (
                    f"=IF(SUM({den_letter}{body_first}:{den_letter}{body_last})>0,"
                    f"SUM({num_letter}{body_first}:{num_letter}{body_last})/"
                    f"SUM({den_letter}{body_first}:{den_letter}{body_last}),\"\")"
                )
                c = ws.cell(row=row, column=ci, value=formula)
            else:
                formula = f"=SUM({col_letter}{body_first}:{col_letter}{body_last})"
                c = ws.cell(row=row, column=ci, value=formula)
            c.alignment = Alignment(horizontal="right", vertical="center")
            if cid in ("amount", "unit_price"):
                c.number_format = '"¥"#,##0'
            elif cid == "cases":
                c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
            else:
                c.number_format = '#,##0'  # 2026-05-28: user 要望 で 全 数値 整数 表示
        else:
            c = ws.cell(row=row, column=ci, value=None)
        c.font = font_total
        c.border = border_all
        c.fill = PatternFill(start_color=SUMMARY_TOTAL_FILL, end_color=SUMMARY_TOTAL_FILL, fill_type="solid")
    ws.row_dimensions[row].height = 22


def _group_inventory_entries(
    entries: list[dict],
) -> list[tuple[tuple[str | None, str | None], list[dict]]]:
    """棚卸エントリ を (大分類, 小分類) で groupby し、 並びを保持して返す。

    entries は 既に (cmaj, cmin, origin, spec, sub_spec) で ソート済み 想定
    (main.py の SQL で ORDER BY 済)。 並び順を 維持しつつ 連続する 同キー を
    まとめる (itertools.groupby 相当)。
    """
    out: list[tuple[tuple[str | None, str | None], list[dict]]] = []
    current_key: tuple[str | None, str | None] | None = None
    current_rows: list[dict] = []
    for r in entries:
        k = (r.get("category_major"), r.get("category_minor"))
        if k != current_key:
            if current_key is not None:
                out.append((current_key, current_rows))
            current_key = k
            current_rows = []
        current_rows.append(r)
    if current_key is not None:
        out.append((current_key, current_rows))
    return out


def _aggregate_stock_by_origin_spec(lots: list[dict], crop_id: int | None = None) -> list[dict]:
    """在庫ロット を (産地, 規格) で 集計。 月末在庫 (end_kg) ベース。

    集計対象: end_kg > 0 のロットのみ (月末在庫が ある分)。
    キー : (origin_name, _format_spec(lot, crop_id))
      crop_id=3 (長芋) は spec_type 単独 で groupby (= AB込み 2L/3L/4L 等 サイズ違い も
      1 区分 に まとめる)、 他 crop は 従来 通り spec+grade+size 連結。
    集計値:
      cases    = Σ (end_kg / kg_per_case)
      total_kg = Σ end_kg
      amount   = Σ (unit_price × end_kg)
    """
    groups: dict[tuple[str, str], dict] = {}
    for lot in lots:
        end_kg = _n(lot.get("end_kg"))
        if end_kg <= 0:
            continue
        origin = (lot.get("origin_name") or "—").strip() or "—"
        spec = _format_spec(lot, crop_id=crop_id)
        key = (origin, spec)
        g = groups.setdefault(key, {
            "origin": origin, "spec": spec,
            "cases": 0.0, "total_kg": 0.0, "amount": 0.0,
        })
        kpc = _n(lot.get("kg_per_case"))
        if kpc > 0:
            g["cases"] += float(end_kg / kpc)
        g["total_kg"] += float(end_kg)
        up = _n(lot.get("unit_price"))
        g["amount"] += float(up * end_kg)
    # 産地 → 規格 順
    return sorted(groups.values(), key=lambda r: (r["origin"], r["spec"]))
