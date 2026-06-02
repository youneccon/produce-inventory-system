"""
api/services/nr_report.py
==========================
NR 原材料使用計算 レポート の 集計ロジック と Excel 生成。

入力: 「商品期間集計」 Excel (列構造):
  D列 = 商品コード, J列 = 合計 kg (= 販売数量 × 重量Kg/個)

処理:
  1. 各行 の (商品コード, 合計kg) を 取得
  2. product_bom と 突合 (商品コード で 完全一致)
  3. 配合展開:
       原料1 kg = 合計 × ratio_1 / 100
       原料2 kg = 合計 × ratio_2 / 100 (二原料商品 のみ)
  4. (origin_id, raw_grade_id) ごと に 合計 kg を 集計
  5. 未登録商品 / 未解決BOM は 警告リスト に 別途出力

出力:
  ExpansionResult:
    - rows: 産地 × 原料規格 × 使用量 kg
    - warnings: 未登録 / 未解決 商品 リスト
    - meta: 入力行数 / 処理済 / スキップ
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import BinaryIO

import psycopg
from psycopg.rows import dict_row
from openpyxl import load_workbook

from api.services.grade_label import compact_grade_label


# =============================================================================
# 入力 列 設定 (商品期間集計 .xlsx)
# =============================================================================
DEFAULT_COL_CODE = 4   # D: 商品コード
DEFAULT_COL_QTY  = 8   # H: 販売数量
DEFAULT_COL_W    = 9   # I: 重量Kg/個
DEFAULT_COL_TOT  = 10  # J: 合計 kg


# =============================================================================
# データ クラス
# =============================================================================
@dataclass
class AggregatedRow:
    """1 (origin, raw_grade) 行 = 集計結果"""
    origin_id: int | None
    origin_text: str           # 表示用 (origin_id 無くても 原文)
    raw_grade_id: int | None
    raw_grade_label: str       # 表示用
    total_kg: Decimal = Decimal(0)


@dataclass
class WarningRow:
    """未登録 / 未解決 商品"""
    excel_row: int
    code: str
    name: str | None
    total_kg: Decimal
    reason: str                # 'not_in_bom' / 'unresolved_bom'


@dataclass
class ExpansionMeta:
    input_rows: int            # 入力行数 (header 除く)
    processed_rows: int        # BOM 突合 成功 行数
    warning_rows: int          # 未登録 + 未解決 行数
    grand_total_kg: Decimal    # 全 合計 (確認用)


@dataclass
class ExpansionResult:
    meta: ExpansionMeta
    rows: list[AggregatedRow] = field(default_factory=list)
    warnings: list[WarningRow] = field(default_factory=list)


def _normalize_code(s) -> str:
    """商品コード 正規化: 10 桁 ゼロ詰め + 大文字"""
    if s is None:
        return ''
    return ('0000000000' + str(s).strip())[-10:].upper()


# =============================================================================
# 取込 & 集計
# =============================================================================
async def expand_from_xlsx(
    conn: psycopg.AsyncConnection,
    file_stream: BinaryIO,
    *,
    col_code: int = DEFAULT_COL_CODE,
    col_total: int = DEFAULT_COL_TOT,
    crop_id: int = 2,
    header_rows: int = 1,
) -> ExpansionResult:
    """商品期間集計 Excel → BOM 展開 → 集計結果。"""
    # Excel 読込
    wb = load_workbook(file_stream, data_only=True, read_only=True)
    ws = wb.active

    input_records: list[tuple[int, str, str, Decimal]] = []   # (row, code, name, total_kg)
    seen_codes: set[str] = set()
    grand_total = Decimal(0)

    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if ri <= header_rows:
            continue
        if not row:
            continue
        code_raw = row[col_code - 1] if col_code - 1 < len(row) else None
        if not code_raw:
            continue
        code = _normalize_code(code_raw)
        if not code or code == '0' * 10:
            continue
        tot_raw = row[col_total - 1] if col_total - 1 < len(row) else None
        if tot_raw is None or tot_raw == '':
            continue
        try:
            total_kg = Decimal(str(tot_raw))
        except Exception:
            continue
        # name は 品名1 (col E = 5)
        name_raw = row[4] if len(row) >= 5 else None
        name = str(name_raw) if name_raw else None
        input_records.append((ri, code, name, total_kg))
        seen_codes.add(code)
        grand_total += total_kg

    # BOM 一括取得 (= 入力 商品コード だけ)
    async with conn.cursor() as cur:
        if seen_codes:
            await cur.execute("""
                SELECT pb.product_code, pb.product_name, pb.is_resolved,
                       pb.origin_id, pb.origin_text,
                       pb.grade_id_1, pb.grade_text_1, pb.ratio_1,
                       pb.grade_id_2, pb.grade_text_2, pb.ratio_2,
                       o.name AS origin_name,
                       g1.spec_type AS g1_spec, g1.grade_level AS g1_grade, g1.size_label AS g1_size,
                       g2.spec_type AS g2_spec, g2.grade_level AS g2_grade, g2.size_label AS g2_size
                FROM product_bom pb
                LEFT JOIN origins o  ON o.id  = pb.origin_id
                LEFT JOIN grades  g1 ON g1.id = pb.grade_id_1
                LEFT JOIN grades  g2 ON g2.id = pb.grade_id_2
                WHERE pb.product_code = ANY(%s)
                  AND pb.crop_id = %s
            """, (list(seen_codes), crop_id))
            bom_map = {r['product_code']: r for r in await cur.fetchall()}
        else:
            bom_map = {}

    # 集計
    agg: dict[tuple[int | None, int | None], AggregatedRow] = {}
    warnings: list[WarningRow] = []
    processed = 0

    def add(origin_id, origin_text, grade_id, grade_label, kg):
        key = (origin_id, grade_id)
        if key not in agg:
            agg[key] = AggregatedRow(
                origin_id=origin_id, origin_text=origin_text or '',
                raw_grade_id=grade_id, raw_grade_label=grade_label or '',
            )
        agg[key].total_kg += kg

    for excel_row, code, name, total_kg in input_records:
        bom = bom_map.get(code)
        if not bom:
            warnings.append(WarningRow(excel_row, code, name, total_kg, 'not_in_bom'))
            continue
        if not bom['is_resolved']:
            warnings.append(WarningRow(excel_row, code, name, total_kg, 'unresolved_bom'))
            # 部分マッチでも 表示可: origin_id ある なら ratio_1 だけでも 集計してもいい
            # → 厳密に 「resolved のみ 集計」 とする (= ユーザー要件 「マスタ整合性 重視」)
            continue
        # 原料1
        ratio_1 = Decimal(bom['ratio_1'])
        kg1 = (total_kg * ratio_1 / Decimal(100)).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
        label_1 = compact_grade_label(bom['g1_spec'], bom['g1_grade'], bom['g1_size'])
        add(bom['origin_id'], bom['origin_name'], bom['grade_id_1'], label_1, kg1)
        # 原料2
        if bom['grade_id_2'] is not None and bom['ratio_2'] is not None:
            ratio_2 = Decimal(bom['ratio_2'])
            kg2 = (total_kg * ratio_2 / Decimal(100)).quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
            label_2 = compact_grade_label(bom['g2_spec'], bom['g2_grade'], bom['g2_size'])
            add(bom['origin_id'], bom['origin_name'], bom['grade_id_2'], label_2, kg2)
        processed += 1

    # 並び順: 産地 → 規格
    sorted_rows = sorted(agg.values(), key=lambda r: (r.origin_text, r.raw_grade_label))

    return ExpansionResult(
        meta=ExpansionMeta(
            input_rows=len(input_records),
            processed_rows=processed,
            warning_rows=len(warnings),
            grand_total_kg=grand_total,
        ),
        rows=sorted_rows,
        warnings=warnings,
    )


# =============================================================================
# Excel 生成 (合計表)
# =============================================================================
def build_summary_xlsx(result: ExpansionResult, *, title: str = '原材料使用計算') -> bytes:
    """ExpansionResult から 合計表 Excel を 生成。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '合計表'

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(border_style='thin', color='888888')
    br = Border(top=thin, bottom=thin, left=thin, right=thin)
    fill_hdr = PatternFill('solid', fgColor='E0F0FF')
    fill_tot = PatternFill('solid', fgColor='FFFBE6')

    # Row 1: タイトル
    ws.cell(1, 1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Row 2: メタ
    m = result.meta
    ws.cell(2, 1, value=f"入力 {m.input_rows} 行 / 処理 {m.processed_rows} / 警告 {m.warning_rows} / 全合計 {float(m.grand_total_kg):,.1f} kg")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    # Row 4: ヘッダ
    headers = ['#', '産地', '原料規格', '使用量 (kg)']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(4, i, value=h)
        c.font = bold; c.fill = fill_hdr; c.alignment = Alignment(horizontal='center')
        c.border = br

    # データ行
    row_idx = 5
    for i, r in enumerate(result.rows, start=1):
        ws.cell(row_idx, 1, value=i).border = br
        ws.cell(row_idx, 2, value=r.origin_text).border = br
        ws.cell(row_idx, 3, value=r.raw_grade_label).border = br
        c = ws.cell(row_idx, 4, value=float(r.total_kg))
        c.border = br; c.number_format = '#,##0.0'; c.alignment = Alignment(horizontal='right')
        row_idx += 1

    # 合計行
    last_data_row = row_idx - 1
    ws.cell(row_idx, 2, value='合計').font = bold
    ws.cell(row_idx, 4, value=f'=SUM(D5:D{last_data_row})')
    for col in range(1, 5):
        c = ws.cell(row_idx, col); c.border = br; c.fill = fill_tot; c.font = bold
        if col == 4:
            c.number_format = '#,##0.0'; c.alignment = Alignment(horizontal='right')

    # 警告 (別シート)
    if result.warnings:
        ws2 = wb.create_sheet('警告')
        ws2.cell(1, 1, value='警告 (BOM 未登録 / 未解決 商品)').font = bold
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        wh = ['行', '商品コード', '品名', '合計 kg', '理由']
        for i, h in enumerate(wh, start=1):
            c = ws2.cell(3, i, value=h); c.font = bold; c.fill = fill_hdr; c.border = br
        for i, w in enumerate(result.warnings, start=4):
            ws2.cell(i, 1, value=w.excel_row)
            ws2.cell(i, 2, value=w.code)
            ws2.cell(i, 3, value=w.name)
            ws2.cell(i, 4, value=float(w.total_kg)).number_format = '#,##0.0'
            reason = '未登録 (BOM に 無い)' if w.reason == 'not_in_bom' else 'マスタ未解決 (origin/grade 未マッピング)'
            ws2.cell(i, 5, value=reason)
            for col in range(1, 6): ws2.cell(i, col).border = br
        for col_letter, width in [('A', 8), ('B', 14), ('C', 30), ('D', 12), ('E', 30)]:
            ws2.column_dimensions[col_letter].width = width

    # 列幅
    for col_letter, width in [('A', 6), ('B', 14), ('C', 22), ('D', 14)]:
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
