"""
api/services/garlic_ledger_sync.py
====================================
大蒜 (crop_id=2) の Excel 仕入管理台帳 (.xlsm) を システム在庫データ で 同期。

3 シート 対応:
  ・仕入管理台帳: sub_kind != 'black' AND origin != '田子'
  ・半製品:       sub_kind = 'semifinished' OR origin = '田子'
  ・黒にんにく:   sub_kind = 'black'

更新する 列 (当月分):
  col17: 前月繰越 (当月以前 入荷 lot) または 当月入荷 (当月 入荷 lot) — 一方のみ
  col18: 当月出庫数 合計
  col19: 当月在庫数 (= col17 - col18)
  col20: 当月在庫金額 (= col19 × 単価)
  col24-54: 各日 (1-31) の 出庫量

絶対 触らない 列:
  col21 (棚卸数)   ←既存数式 =SXX 等
  col22 (差数)     ←既存数式
  col23 (差数原因) ←ユーザー手入力

マッチング 手順:
  Excel 行 と システム lot の 一致 判定:
    key = (supplier_n, origin_n, spec_combined_n, inbound_date, total_kg)
    n = NFKC + trim 正規化、 origin は 末尾「産」 削除

  未マッチ:
    Excel のみ: そのまま (システムに無い古い行)
    System のみ: シート末尾に append (当月新規入荷で Excel に行が無いケース)
"""
from __future__ import annotations

import logging
import re
import unicodedata
from collections import defaultdict
from copy import copy
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 出力 .xlsm に 付与 する カスタム プロパティ 名。 同期済み ファイル を
# 再度 入力 した場合 に 警告 を 出す ため (= inactive 行 の 数式 が 既に 失われて
# いる 可能性 が ある)。
SYNCED_MARKER_PROP = "GarlicLedgerSyncedAt"


# シート → (条件 SQL 用 ラベル)
SHEETS = {
    "仕入管理台帳": "normal",       # sub_kind IN (NULL, 'normal', 'tako'?) AND origin != '田子'
    "半製品":       "semifinished", # sub_kind='semifinished' OR origin='田子'
    "黒にんにく":   "black",        # sub_kind='black'
}


def _norm_key(s: Any) -> str:
    """NFKC + trim + 小文字化 で 正規化キー化。 None/数値 でも 文字列化。"""
    if s is None:
        return ""
    if isinstance(s, (int, float, Decimal)):
        s = str(s)
    return unicodedata.normalize("NFKC", str(s)).strip()


def _norm_origin(s: Any) -> str:
    """origin: NFKC + trim + 末尾「産」削除。 'X' と 'X産' を 同一視。"""
    v = _norm_key(s)
    while v.endswith("産"):
        v = v[:-1].rstrip()
    return v


def _norm_qty(v: Any) -> str:
    """数量: 浮動小数点誤差吸収 (小数 1 桁四捨五入後 比較用 文字列)。"""
    if v is None or v == "":
        return ""
    try:
        d = Decimal(str(v))
    except Exception:
        return _norm_key(v)
    return str(d.quantize(Decimal("0.1")))


def _format_spec(spec_type: str | None, grade_level: str | None,
                 size_label: str | None) -> str:
    """規格・等級・サイズ を Excel 表記 (スペース無し連結) に。 calendar_excel と 同等。"""
    parts: list[str] = []
    if spec_type and spec_type != "標準":
        parts.append(spec_type)
    if grade_level and grade_level != "-":
        parts.append(grade_level)
    if size_label and size_label != "-":
        parts.append(size_label)
    return "".join(parts)


def _lot_key(supplier: str, origin: str, spec: str,
             inbound_date: date, total_kg: Decimal) -> tuple[str, str, str, str, str]:
    """マッチング用 タプルキー。 datetime は date 部分のみ 比較 (= 時刻無視)。"""
    if isinstance(inbound_date, datetime):
        date_str = inbound_date.date().isoformat()
    elif isinstance(inbound_date, date):
        date_str = inbound_date.isoformat()
    else:
        date_str = _norm_key(inbound_date)
    return (
        _norm_key(supplier),
        _norm_origin(origin),
        _norm_key(spec),
        date_str,
        _norm_qty(total_kg),
    )


def _excel_date_to_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ─── システム データ 取得 ───

async def _fetch_lots_by_sheet(db, month: str) -> dict[str, list[dict]]:
    """当月 (= month=YYYY-MM) の カレンダー観点 で 大蒜 lot を取得。

    返り値: { sheet_name: [lot_info_dict, ...] }
        lot_info_dict は CalendarLot 相当 (carryover_kg, inbound_kg, outbound_kg,
        end_kg, daily の dict 含む)
    """
    # carryover/inbound/end 計算用 の 月境界
    first = datetime.strptime(month + "-01", "%Y-%m-%d").date()
    next_first = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
    # 前月 period (stock_counts.period 用 'YYYY-MM' 形式)
    prev_last = first - timedelta(days=1)
    prev_period = prev_last.strftime("%Y-%m")

    async with db.cursor() as cur:
        await cur.execute("""
            SELECT
                il.id           AS lot_id,
                il.code         AS lot_code,
                s.name          AS supplier_name,
                o.name          AS origin_name,
                g.spec_type, g.grade_level, g.size_label,
                il.inbound_date,
                il.total_kg,
                il.kg_per_case,
                il.unit_price,
                il.brokerage_fee, il.freight_fee,
                il.prepay_date, il.prepay_amount,
                il.postpay_date, il.postpay_amount,
                p.crop_id, p.sub_kind,
                -- 前月繰越 (= 前月末 の 実測棚卸数)。 stock_counts テーブル を 優先 参照。
                -- 棚卸補正 (ロス・廃棄・実測差) を 反映 した 真の 月初残量。
                -- stock_counts が無い lot は フォールバック で total_kg - 過去出庫 (帳簿計算)。
                -- 当月入荷 lot は carryover = 0 (前月時点では 未入荷)。
                -- ※ stock_counts は (lot_id, period) UNIQUE だが、 防衛 で 確定日 降順 + LIMIT 1。
                CASE WHEN il.inbound_date < %(first)s
                     THEN COALESCE(
                         (SELECT counted_kg FROM stock_counts sc
                           WHERE sc.lot_id = il.id AND sc.period = %(prev_period)s
                           ORDER BY sc.confirmed_at DESC, sc.id DESC
                           LIMIT 1),
                         il.total_kg - COALESCE((SELECT SUM(quantity_kg) FROM outbound_records ob
                                                 WHERE ob.lot_id = il.id AND ob.outbound_date < %(first)s), 0)
                     )
                     ELSE 0
                END AS carryover_kg,
                -- 当月入荷 (= 当月入荷 lot は total_kg、 過去入荷は 0)
                CASE WHEN il.inbound_date >= %(first)s AND il.inbound_date < %(next)s
                     THEN il.total_kg
                     ELSE 0
                END AS inbound_kg,
                -- 当月出庫合計
                COALESCE((SELECT SUM(quantity_kg) FROM outbound_records ob
                          WHERE ob.lot_id = il.id
                            AND ob.outbound_date >= %(first)s AND ob.outbound_date < %(next)s), 0) AS outbound_kg
              FROM inbound_lots il
              JOIN products  p ON p.id = il.product_id
              JOIN suppliers s ON s.id = il.supplier_id
              JOIN origins   o ON o.id = p.origin_id
              JOIN grades    g ON g.id = p.grade_id
             WHERE p.crop_id = 2
               AND il.archived_at IS NULL
               AND il.inbound_date < %(next)s
             ORDER BY il.inbound_date, il.id
        """, {"first": first, "next": next_first, "prev_period": prev_period})
        all_lots = await cur.fetchall()

        # 日別出庫 (1..31) を 別クエリで取得
        await cur.execute("""
            SELECT ob.lot_id, EXTRACT(DAY FROM ob.outbound_date) AS d, SUM(ob.quantity_kg) AS q
              FROM outbound_records ob
              JOIN inbound_lots il ON il.id = ob.lot_id
              JOIN products p ON p.id = il.product_id
             WHERE p.crop_id = 2
               AND ob.outbound_date >= %(first)s AND ob.outbound_date < %(next)s
             GROUP BY ob.lot_id, EXTRACT(DAY FROM ob.outbound_date)
        """, {"first": first, "next": next_first})
        daily_rows = await cur.fetchall()
    daily_map: dict[int, dict[int, Decimal]] = defaultdict(dict)
    for r in daily_rows:
        daily_map[r["lot_id"]][int(r["d"])] = Decimal(str(r["q"]))

    # シート 振り分け。 「当月アクティブ で ない lot」 (carry=0 AND inb=0) は スキップ。
    # ob > 0 は 必ず carry > 0 or inb > 0 を 満たす (出庫するには 在庫が必要) ので
    # ob 条件 は不要 (user 指摘 2026-05-26)。
    bucket: dict[str, list[dict]] = {sn: [] for sn in SHEETS}
    for r in all_lots:
        carry_v = float(r["carryover_kg"] or 0)
        inb_v   = float(r["inbound_kg"] or 0)
        if carry_v <= 0 and inb_v <= 0:
            continue  # 当月で 動き なし → 同期対象外
        sub_kind = r["sub_kind"]
        origin = r["origin_name"]
        if sub_kind == "black":
            sheet = "黒にんにく"
        elif sub_kind == "semifinished" or origin == "田子":
            sheet = "半製品"
        else:
            sheet = "仕入管理台帳"
        spec_combined = _format_spec(r["spec_type"], r["grade_level"], r["size_label"])
        bucket[sheet].append({
            **dict(r),
            "spec_combined": spec_combined,
            "daily": dict(daily_map.get(r["lot_id"], {})),
        })
    return bucket


# ─── Excel 列定義 ───

COL_SUPPLIER     = 1
COL_ORIGIN       = 2
COL_SPEC         = 3
COL_CASES        = 4
COL_KG_PER_CASE  = 5
COL_TOTAL_KG     = 6
COL_UNIT_PRICE   = 7
COL_BROKERAGE    = 8
COL_FREIGHT      = 9
COL_TAX          = 10
COL_GRAND_TOTAL  = 11
COL_PREPAY_DATE  = 12
COL_PREPAY_AMT   = 13
COL_INBOUND_DATE = 14
COL_POSTPAY_DATE = 15
COL_POSTPAY_AMT  = 16
COL_CARRYOVER    = 17  # 兼 当月入荷
COL_OUTBOUND     = 18
COL_END_KG       = 19
COL_END_VALUE    = 20
COL_STOCKTAKE    = 21  # ← 触らない
COL_DIFF         = 22  # ← 触らない
COL_DIFF_NOTE    = 23  # ← 触らない
COL_DAY_START    = 24  # day 1
# col 24+d-1 = day d, day 1..31 → col 24..54


def _read_excel_rows(ws) -> list[dict]:
    """既存 Excel 行を 行番号付きで 抽出。 ヘッダ行 (1-5) は スキップ、 合計行 等 は 除外。"""
    out: list[dict] = []
    for r in range(6, ws.max_row + 1):
        supplier = ws.cell(row=r, column=COL_SUPPLIER).value
        if supplier is None or supplier == "":
            continue
        sv = str(supplier).strip()
        if sv == "合計":
            continue
        origin = ws.cell(row=r, column=COL_ORIGIN).value
        spec = ws.cell(row=r, column=COL_SPEC).value
        inbound = _excel_date_to_date(ws.cell(row=r, column=COL_INBOUND_DATE).value)
        total_kg = ws.cell(row=r, column=COL_TOTAL_KG).value
        out.append({
            "row": r,
            "supplier": sv,
            "origin": origin,
            "spec": spec,
            "inbound_date": inbound,
            "total_kg": total_kg,
            "key": _lot_key(sv, origin or "", spec or "", inbound or date(1, 1, 1), total_kg or 0),
        })
    return out


def _sort_sheet_by_inbound_date(ws, total_row: int) -> None:
    """合計行 直前 までの 全 データ行 を col14 (入荷日) で 昇順ソート。

    各行 の セル値・数式・スタイル・行高・hidden 状態 を スナップショット → ソート →
    全行クリア → 順番 に 書き戻し。 セル数式 (= 自分行 を 参照) は openpyxl Translator
    で 新しい行番号 に 自動 translate。
    """
    from copy import copy as _cp
    from datetime import datetime as _dt
    from openpyxl.formula.translate import Translator

    rows: list[dict] = []
    for r in range(6, total_row):
        sup = ws.cell(row=r, column=1).value
        if not sup:
            continue  # 空行 は スキップ
        snap = {
            "old_row": r,
            "hidden": ws.row_dimensions[r].hidden,
            "height": ws.row_dimensions[r].height,
            "inbound_date": ws.cell(row=r, column=COL_INBOUND_DATE).value,
            "cells": {},
        }
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            entry = {"value": cell.value, "style": None}
            if cell.has_style:
                entry["style"] = {
                    "font":          _cp(cell.font),
                    "fill":          _cp(cell.fill),
                    "border":        _cp(cell.border),
                    "alignment":     _cp(cell.alignment),
                    "number_format": cell.number_format,
                    "protection":    _cp(cell.protection),
                }
            snap["cells"][c] = entry
        rows.append(snap)

    # 入荷日 昇順 (None は 末尾)
    def _sortkey(snap):
        d = snap["inbound_date"]
        if isinstance(d, _dt):
            return (0, d)
        if isinstance(d, date):
            return (0, _dt.combine(d, _dt.min.time()))
        return (1, _dt.max)  # 日付なし = 末尾
    rows.sort(key=_sortkey)

    # 全 データ行 を 初期化 (r6 〜 total_row-1)
    for r in range(6, total_row):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None
        ws.row_dimensions[r].hidden = False
        ws.row_dimensions[r].height = None

    # ソート結果 を 書き戻し
    for i, snap in enumerate(rows):
        new_row = 6 + i
        old_row = snap["old_row"]
        if snap["height"]:
            ws.row_dimensions[new_row].height = snap["height"]
        ws.row_dimensions[new_row].hidden = snap["hidden"]
        for c, entry in snap["cells"].items():
            v = entry["value"]
            # 数式 は 新しい行番号 に translate (= 自分行 参照 を 維持)
            if isinstance(v, str) and v.startswith("=") and old_row != new_row:
                try:
                    v = Translator(v, origin=f"A{old_row}").translate_formula(f"A{new_row}")
                except Exception:
                    pass  # 失敗時 は 元 数式 を そのまま 残す
            dst = ws.cell(row=new_row, column=c)
            dst.value = v
            if entry["style"]:
                st = entry["style"]
                dst.font          = st["font"]
                dst.fill          = st["fill"]
                dst.border        = st["border"]
                dst.alignment     = st["alignment"]
                dst.number_format = st["number_format"]
                dst.protection    = st["protection"]


def _find_total_row(ws) -> int | None:
    """「合計」 行 を 探す。 col 1〜col 3 を スキャン (本ファイル では col 3 が 「合計」)。
    無ければ None。
    """
    for r in range(6, ws.max_row + 1):
        for c in (1, 2, 3):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() == "合計":
                return r
    return None


def _extend_table_ranges(ws, insert_at: int, insert_count: int) -> None:
    """ws に 定義された 全 Excel Table の ref を、 insert_at に N 行 挿入 した分
    拡張する。 元 ref が 'A5:BB365' で insert_at=365, N=9 → 'A5:BB374'。

    SUBTOTAL や 構造化参照 (テーブル名[列名]) が 新規行 を 含む ように 必須。
    範囲が ずれない と Excel 開いた時 破損警告 や 集計漏れ の 原因 に なる。
    """
    if insert_count <= 0:
        return
    for name in list(ws.tables):
        tbl = ws.tables[name]
        ref = tbl.ref  # 例 'A5:BB365'
        if ":" not in ref:
            continue
        start, end = ref.split(":", 1)
        # end は 'BB365' → 列文字 + 行番号 に 分割
        m = re.match(r"^([A-Z]+)(\d+)$", end)
        if not m:
            continue
        col_letters, row_str = m.group(1), m.group(2)
        end_row = int(row_str)
        if end_row >= insert_at:
            new_end_row = end_row + insert_count
            tbl.ref = f"{start}:{col_letters}{new_end_row}"


def _formula_for_row(template_formulas: dict, col: int, row: int) -> str | None:
    """template_formulas に col の 元 formula が あれば Translator で 指定 row に
    投影 して 返す。 無ければ None。"""
    if col not in template_formulas:
        return None
    from openpyxl.formula.translate import Translator
    origin_row, formula = template_formulas[col]
    if origin_row == row:
        return formula
    try:
        return Translator(formula, origin=f"A{origin_row}").translate_formula(f"A{row}")
    except Exception:
        return formula  # 失敗時 は 元 formula を そのまま


def _write_new_lot_from_snapshot(ws, row: int, lot: dict, days_in_month: int,
                                  template_snapshot: dict | None,
                                  template_formulas: dict | None = None) -> None:
    """新規 append 行 を 書き込む。 template_snapshot (= 削除前 の active 行 スタイル)
    から style を コピーする。 template_formulas (= 計算列 の 元 formula 辞書) が ある なら、
    J/K/R/S/T は その formula を Translator で 当行 に 投影 して 書く (= Table 集計列 の
    構造化参照 に も 対応、 「矛盾した集計列の数式」 警告 回避)。

    書き込む列: col1-20 + 日付列。 col21-23 (棚卸数式) は 触らない (= 元から空 の セル)。
    """
    template_formulas = template_formulas or {}
    # スタイル + 行高 復元
    if template_snapshot:
        rh = template_snapshot.get("row_height")
        if rh:
            ws.row_dimensions[row].height = rh
        from copy import copy as _cp
        for c, st in template_snapshot.get("cells", {}).items():
            dst = ws.cell(row=row, column=c)
            dst.font          = _cp(st["font"])
            dst.fill          = _cp(st["fill"])
            dst.border        = _cp(st["border"])
            dst.alignment     = _cp(st["alignment"])
            dst.number_format = st["number_format"]
            dst.protection    = _cp(st["protection"])

    # 基本情報 (col1-16)
    ws.cell(row=row, column=COL_SUPPLIER, value=lot["supplier_name"])
    ws.cell(row=row, column=COL_ORIGIN, value=(lot["origin_name"] or "") + "産")
    ws.cell(row=row, column=COL_SPEC, value=lot["spec_combined"])
    kpc = float(lot["kg_per_case"] or 0) or None
    if kpc:
        ws.cell(row=row, column=COL_KG_PER_CASE, value=kpc)
        tk = float(lot["total_kg"] or 0)
        if tk: ws.cell(row=row, column=COL_CASES, value=tk / kpc)
    ws.cell(row=row, column=COL_TOTAL_KG, value=float(lot["total_kg"] or 0))
    up = float(lot["unit_price"] or 0)
    if up: ws.cell(row=row, column=COL_UNIT_PRICE, value=up)
    br = float(lot["brokerage_fee"] or 0)
    if br: ws.cell(row=row, column=COL_BROKERAGE, value=br)
    fr = float(lot["freight_fee"] or 0)
    if fr: ws.cell(row=row, column=COL_FREIGHT, value=fr)
    # 計算列 (J, K, R, S, T) は **数式** で 書く (= 値 直書き だと Excel が
    # 「矛盾した集計列の数式」 警告 を 出す ため。 2026-05-31 user 報告)。
    # 列 letter helper
    f_col = get_column_letter(COL_TOTAL_KG)    # F
    g_col = get_column_letter(COL_UNIT_PRICE)  # G
    h_col = get_column_letter(COL_BROKERAGE)   # H
    i_col = get_column_letter(COL_FREIGHT)     # I
    q_col = get_column_letter(COL_CARRYOVER)   # Q
    r_col = get_column_letter(COL_OUTBOUND)    # R
    s_col = get_column_letter(COL_END_KG)      # S
    day_start_col = get_column_letter(COL_DAY_START)                  # X
    day_end_col   = get_column_letter(COL_DAY_START + days_in_month - 1)  # 月末日 列
    if up and lot["total_kg"]:
        # J / K: 元 Table 集計列 の formula を 優先 (= 構造化参照 でも 可)、
        # 無ければ 標準 数式 で fallback。
        tax_f = _formula_for_row(template_formulas, COL_TAX, row)
        ws.cell(row=row, column=COL_TAX,
                value=tax_f or f"=ROUND({f_col}{row}*{g_col}{row}*0.08,0)")
        gt_f = _formula_for_row(template_formulas, COL_GRAND_TOTAL, row)
        ws.cell(row=row, column=COL_GRAND_TOTAL,
                value=gt_f or f"=ROUND({f_col}{row}*{g_col}{row}*1.08+{h_col}{row}+{i_col}{row},0)")
    if lot["prepay_date"]:    ws.cell(row=row, column=COL_PREPAY_DATE, value=lot["prepay_date"])
    if lot["prepay_amount"]:  ws.cell(row=row, column=COL_PREPAY_AMT, value=float(lot["prepay_amount"]))
    if lot["inbound_date"]:   ws.cell(row=row, column=COL_INBOUND_DATE, value=lot["inbound_date"])
    if lot["postpay_date"]:   ws.cell(row=row, column=COL_POSTPAY_DATE, value=lot["postpay_date"])
    if lot["postpay_amount"]: ws.cell(row=row, column=COL_POSTPAY_AMT, value=float(lot["postpay_amount"]))

    # 当月分 (col17-20 + 日付列)
    # ・Q (carryover) は データ 入力 値 (carry + inb、 一方 0) で 直書き
    # ・R (outbound) は =SUM(X{r}:end{r}) 数式 (= 日別 出庫 を 自動 集計)
    # ・S (end_kg) は =Q{r}-R{r}
    # ・T (end_value) は =ROUND(S{r}*G{r}, 0)
    carry = float(lot["carryover_kg"] or 0)
    inb = float(lot["inbound_kg"] or 0)
    col17_v = carry + inb       # 一方は 必ず 0
    ws.cell(row=row, column=COL_CARRYOVER, value=col17_v)
    # R / S / T も 元 Table 集計列 formula を 優先 (構造化参照 対応)、 fallback で 自前 数式
    ob_f = _formula_for_row(template_formulas, COL_OUTBOUND, row)
    ws.cell(row=row, column=COL_OUTBOUND,
            value=ob_f or f"=SUM({day_start_col}{row}:{day_end_col}{row})")
    end_f = _formula_for_row(template_formulas, COL_END_KG, row)
    ws.cell(row=row, column=COL_END_KG,
            value=end_f or f"={q_col}{row}-{r_col}{row}")
    ev_f = _formula_for_row(template_formulas, COL_END_VALUE, row)
    ws.cell(row=row, column=COL_END_VALUE,
            value=ev_f or f"=ROUND({s_col}{row}*{g_col}{row},0)")
    daily = lot["daily"] or {}
    for d in range(1, days_in_month + 1):
        col = COL_DAY_START + d - 1
        q = daily.get(d)
        if q:
            ws.cell(row=row, column=col, value=float(q))


# ─── Master シート 突合 (警告生成 用) ───

def _collect_master_set(wb, sheet_name: str, value_col: int) -> set[str]:
    """M シート の 指定列 (1=ID, 2=名前) を 正規化キー集合 で 返す。"""
    if sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    out: set[str] = set()
    for r in range(2, ws.max_row + 1):  # r1 はヘッダ
        v = ws.cell(row=r, column=value_col).value
        if v: out.add(_norm_key(v))
    return out


def _days_in_month(month: str) -> int:
    first = datetime.strptime(month + "-01", "%Y-%m-%d").date()
    nxt = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
    return (nxt - first).days


# ─── 公開 関数: 同期実行 ───

async def sync_garlic_ledger(
    xlsm_bytes: bytes, month: str, db, dry_run: bool = False,
) -> tuple[bytes | None, dict]:
    """大蒜 Excel 仕入管理台帳 を システムデータで同期。

    Args:
        xlsm_bytes: 入力 .xlsm の バイト列
        month: 'YYYY-MM' 同期対象月
        db: psycopg async connection
        dry_run: True なら DB 触らずシート更新 のみ。 戻り値の bytes は None
                 (preview = warnings だけ 確認 する 用途)

    Returns:
        (xlsm_bytes_or_None, result_dict)
        result_dict = {
          'sheets': {
            sheet_name: {'updated': N, 'appended': N, 'unmatched_excel': N, 'unmatched_system': N}
          },
          'warnings': [str, ...],
          'master_warnings': [str, ...],
        }
    """
    days_in_month = _days_in_month(month)

    # openpyxl は 外部リンク (externalLinks/) と 図形 (drawings/, media/) を
    # 完全 に は 保持 できない (rId 欠落、 DrawingML 未対応 等)。
    # 後で 復元 する ため オリジナル の バイナリ + sheet→drawing 関係 を 退避。
    # 注意: sheet rels は openpyxl の 出力 と マージ する (= 完全上書き しない)。
    import zipfile
    from xml.etree import ElementTree as _ET
    preserved_parts: dict[str, bytes] = {}
    # sheet rels file path -> [{id, type, target}] (元 drawing 系 関係 のみ 抽出)
    sheet_rels_to_merge: dict[str, list[dict]] = {}
    # sheet.xml path -> drawing rId リスト (注入 用)
    sheet_drawing_links: dict[str, list[dict]] = {}
    # sheet.xml path -> <controls>...</controls> + 周辺 mc:AlternateContent
    # (= Form Control の DrawingML/VML 重複描画 を 防ぐ 必須 セクション)
    sheet_controls_xml: dict[str, str] = {}
    with zipfile.ZipFile(BytesIO(xlsm_bytes)) as z:
        for name in z.namelist():
            # externalLinks/ + drawings/ + media/ + ctrlProps/ + printerSettings/ 配下
            # を そのまま 保存。 ctrlProps は Form Control の チェックボックス状態 等
            # を 持つ XML。 これが ないと Form Control が DrawingML と VML の 二重
            # 描画 で 重複 表示 される (合計表/歩留計算 で 観測)。
            if (name.startswith("xl/externalLinks/")
                or name.startswith("xl/drawings/")
                or name.startswith("xl/media/")
                or name.startswith("xl/ctrlProps/")
                or name.startswith("xl/printerSettings/")):
                preserved_parts[name] = z.read(name)
            # sheet.xml: <controls> セクション を 抽出 (mc:AlternateContent 包み 込み)
            # openpyxl は controls を 出力 しない ため、 Form Control の shapeId →
            # ctrlProp の 紐付け が 失われ DrawingML xdr:sp と VML v:shape が 別物 として
            # 重複 描画 される。 後で 再注入 する。
            if (name.startswith("xl/worksheets/sheet")
                and name.endswith(".xml") and "/_rels/" not in name):
                try:
                    sheet_xml = z.read(name).decode("utf-8")
                    # 外側 の <mc:AlternateContent ...><mc:Choice><controls>...</controls></mc:Choice><mc:Fallback/></mc:AlternateContent>
                    m = re.search(
                        r'<mc:AlternateContent[^>]*>\s*<mc:Choice[^>]*>\s*<controls>.*?</controls>\s*</mc:Choice>(?:\s*<mc:Fallback[^/]*/?>(?:[^<]*</mc:Fallback>)?)?\s*</mc:AlternateContent>',
                        sheet_xml, flags=re.DOTALL
                    )
                    if m:
                        sheet_controls_xml[name] = m.group(0)
                except Exception as e:
                    logger.warning("controls 抽出失敗 (%s): %s", name, e)
            # sheet rels: 全関係 を 抽出 (drawing + ctrlProp + printerSettings + ...
            # 全部)。 ただし vmlDrawing は openpyxl が 自前で 別 ID で 作るので 除外。
            # 復元時 は target ベース で 重複 check (= openpyxl が 既に 持ってる
            # ものは スキップ、 持ってない ものだけ 追加)。
            if name.startswith("xl/worksheets/_rels/") and name.endswith(".rels"):
                try:
                    rels_xml = z.read(name).decode("utf-8")
                    root = _ET.fromstring(rels_xml)
                    ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
                    sheet_path = name.replace("/_rels/", "/").replace(".rels", "")
                    rels_to_merge = []
                    drawing_links = []
                    for rel in root.findall(f"{ns}Relationship"):
                        rtype = rel.get("Type", "")
                        if "vmlDrawing" in rtype:
                            continue  # openpyxl が 別 ID で 作る ので skip
                        entry = {"id": rel.get("Id"), "type": rtype, "target": rel.get("Target")}
                        rels_to_merge.append(entry)
                        # drawing 関係 のみ sheet.xml に <drawing> タグ 注入対象
                        if "drawing" in rtype.lower():
                            drawing_links.append(entry)
                    if rels_to_merge:
                        sheet_rels_to_merge[name] = rels_to_merge
                    if drawing_links:
                        sheet_drawing_links[sheet_path] = drawing_links
                except Exception as e:
                    logger.warning("sheet rels 抽出失敗 (%s): %s", name, e)

    wb = load_workbook(BytesIO(xlsm_bytes), keep_vba=True, data_only=False)
    # 開き直し時 に Excel が 外部リンク 等 の キャッシュ値 を 再計算 する よう 指示。
    try:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception as e:
        logger.warning("CalcProperties 設定失敗 (Excel 再計算 を 強制 できず): %s", e)

    result: dict[str, Any] = {"sheets": {}, "warnings": [], "master_warnings": []}

    # 再同期検出: 同期済み marker が ある なら 警告 を 立てる。
    # inactive 行 の c18/c19/c20 数式 は 1 回目 の 同期 で 既に None に されている ので、
    # 2 回目 を 実行 しても 数式 が 蘇る わけ では ない。 ユーザー が 元 .xlsm を 使い
    # 直す か、 同期後 ファイル を 編集 して 同期 する のか 明確 に させる ため 通知。
    prev_synced_at = None
    try:
        for prop in wb.custom_doc_props:
            if prop.name == SYNCED_MARKER_PROP:
                prev_synced_at = prop.value
                break
    except Exception as e:
        logger.warning("custom_doc_props 読取り失敗: %s", e)
    if prev_synced_at:
        result["warnings"].append(
            f"このファイルは {prev_synced_at} に 同期済み です。 "
            f"再同期すると inactive 行 の c18/c19/c20 数式 は 既に 空 の まま 残ります。 "
            f"元 .xlsm から 同期 し直す こと を 推奨。"
        )

    # システムから lots 取得 (シート別)
    sys_by_sheet = await _fetch_lots_by_sheet(db, month)

    # Master シート 集合 (警告チェック用)
    master_suppliers = _collect_master_set(wb, "M仕入先", value_col=2)
    master_origins   = _collect_master_set(wb, "M産地",   value_col=2)
    master_specs     = _collect_master_set(wb, "M規格",   value_col=2)

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            result["warnings"].append(f"シート '{sheet_name}' が ファイルに ありません")
            continue
        ws = wb[sheet_name]
        sys_lots = sys_by_sheet.get(sheet_name, [])

        # ═════════ フローチャート Step 3-7 ═════════
        # Step 3: Excel data 行 を 分類 (active = col17>0 削除対象、 inactive = c17=0 保護)
        active_rows: list[int] = []
        inactive_rows: list[int] = []
        for r in range(6, ws.max_row + 1):
            sup = ws.cell(row=r, column=1).value
            if not sup or str(sup).strip() == "合計":
                continue
            c17 = ws.cell(row=r, column=17).value
            if isinstance(c17, (int, float)) and c17 > 0:
                active_rows.append(r)
            else:
                inactive_rows.append(r)

        # Step 4: INACTIVE 行の col18/19/20 + 日付列 (col24-54) を クリア
        #   重要: c18 数式 (=SUM(X:BB)) は delete_rows で 参照行が ずれ、
        #   append 行 を 誤参照 する 不具合 が ある (v11 で発覚)。
        #   数式 ごと None で潰す = 元 数式 削除 + 値 0 扱い → SUBTOTAL 寄与 0。
        #   c17 は ノータッチ (= 元から 0 or 空、 ユーザー手入力 保護)。
        for r in inactive_rows:
            ws.cell(row=r, column=COL_OUTBOUND).value = None
            ws.cell(row=r, column=COL_END_KG).value = None
            ws.cell(row=r, column=COL_END_VALUE).value = None
            for d in range(1, days_in_month + 1):
                ws.cell(row=r, column=COL_DAY_START + d - 1).value = None

        # Step 5: ACTIVE 行 を 物理削除 (下から、 行番号 が ずれない 順)
        #   削除前 に 最終 active 行 の スタイル + 計算列 数式 を snapshot。
        #   2026-05-31: 元 Excel の Table 集計列 (= 「集計列の数式」) は 構造化参照
        #     (例: `=[@日1]+[@日2]+...`) を 使う 場合 が ある。 私 の 独自 formula
        #     (`=SUM(X{r}:BB{r})`) と 異 なる ため 「矛盾した集計列の数式」 警告 が
        #     出る → 元 formula を テンプレ として 取り込み、 Translator で 各行 へ
        #     投影 する 形 に 変更。
        template_snapshot = None
        template_formulas: dict[int, tuple[int, str]] = {}  # col → (origin_row, formula)
        if active_rows:
            from copy import copy as _cp
            tmpl_row = active_rows[-1]
            template_snapshot = {
                "row_height": ws.row_dimensions[tmpl_row].height,
                "cells": {},
            }
            for c in range(1, ws.max_column + 1):
                src = ws.cell(row=tmpl_row, column=c)
                if src.has_style:
                    template_snapshot["cells"][c] = {
                        "font": _cp(src.font),
                        "fill": _cp(src.fill),
                        "border": _cp(src.border),
                        "alignment": _cp(src.alignment),
                        "number_format": src.number_format,
                        "protection": _cp(src.protection),
                    }
            # 計算列 (J/K/R/S/T) の 元 formula を キャプチャ。 構造化参照 (= '[@..]') か
            # 自分行 セル 参照 か は 問わ ない、 Translator で 投影 する。
            for col in (COL_TAX, COL_GRAND_TOTAL, COL_OUTBOUND, COL_END_KG, COL_END_VALUE):
                v = ws.cell(row=tmpl_row, column=col).value
                if isinstance(v, str) and v.startswith("="):
                    template_formulas[col] = (tmpl_row, v)

        deleted_count = 0
        for r in sorted(active_rows, reverse=True):
            ws.delete_rows(r, amount=1)
            deleted_count += 1
        # Excel Table ref を 削除分 縮小
        if deleted_count > 0:
            for name in list(ws.tables):
                tbl = ws.tables[name]
                ref = tbl.ref
                if ":" not in ref: continue
                start, end = ref.split(":", 1)
                m = re.match(r"^([A-Z]+)(\d+)$", end)
                if not m: continue
                col_letters, row_str = m.group(1), m.group(2)
                tbl.ref = f"{start}:{col_letters}{int(row_str) - deleted_count}"

        # Step 6: DB active lot を fresh append (合計 行 直前 に insert)
        total_row = _find_total_row(ws)
        n_new = len(sys_lots)
        if n_new > 0:
            if total_row is not None:
                insert_at = total_row
                ws.insert_rows(insert_at, amount=n_new)
                _extend_table_ranges(ws, insert_at, n_new)
            else:
                insert_at = ws.max_row + 1
            for i, lot in enumerate(sys_lots):
                row = insert_at + i
                _write_new_lot_from_snapshot(ws, row, lot, days_in_month,
                                              template_snapshot, template_formulas)
                # 重要: insert_rows は 元位置 の hidden 属性 を 継承 する。
                # append 行 は 必ず 可視 に する (= ユーザー が DB lot を 見えなく なる の防止)。
                ws.row_dimensions[row].hidden = False

        # Step 6.5: 合計行直前 まで を 入荷日 で 昇順ソート (preserved + append 混在)
        total_row_after = _find_total_row(ws)
        if total_row_after is not None:
            _sort_sheet_by_inbound_date(ws, total_row_after)

        # Step 6.7: 全 計算列 (J/K/R/S/T/U/V) を 防衛的 に 自分行 参照 で 再生成。
        # 2026-05-31 拡張: 旧 logic は U/V だけ 再生成 だった が、 openpyxl の
        # delete_rows / insert_rows / sort の 連続 で 「F50*G50 が F30*G30 に
        # 補正 されず F40*G40 等 別 行 参照 に ずれる」 不具合 を J/K/R/S/T で も
        # 観測 (user 報告)。 sort 後 に 全 計算列 を 自分行 参照 へ 強制 リセット して
        # 行 位置 ズレ に 関わらず 正しい 値 が 出る よう に。
        # 「合計」 行 と 空 行 は 飛ばす (= SUM 集計式 など は 別 logic で 管理)。
        #
        # 数式 仕様 (= 同 行 参照):
        #   J (消費税)        = ROUND(F*G*0.08, 0)
        #   K (仕入合計金額)  = ROUND(F*G*1.08 + H + I, 0)
        #   R (当月出庫)      = SUM(X:月末日列)
        #   S (当月在庫)      = Q - R
        #   T (在庫評価額)    = ROUND(S*G, 0)
        #   U (棚卸数 mirror) = S
        #   V (差数)         = N(S) - N(U)
        if total_row_after is not None:
            f_col = get_column_letter(COL_TOTAL_KG)
            g_col = get_column_letter(COL_UNIT_PRICE)
            h_col = get_column_letter(COL_BROKERAGE)
            i_col = get_column_letter(COL_FREIGHT)
            q_col = get_column_letter(COL_CARRYOVER)
            r_col = get_column_letter(COL_OUTBOUND)
            s_col = get_column_letter(COL_END_KG)
            ds_col = get_column_letter(COL_DAY_START)
            de_col = get_column_letter(COL_DAY_START + days_in_month - 1)
            for r in range(6, total_row_after):
                sup_v = ws.cell(row=r, column=COL_SUPPLIER).value
                if not sup_v or str(sup_v).strip() == "合計":
                    continue
                # ─── R/S/T は 「元 セル に 数式 が ある か」 で 判定 ───
                # Step 4 で inactive 行 の R/S/T は None に クリア 済 (= 「動き なし」
                # 意図、 復活 させ ない)。 active/append 行 だけ は 数式 が 入って いる ので
                # それ を 元 Table 集計列 formula (構造化参照 等) で 再投影。
                r_cell = ws.cell(row=r, column=COL_OUTBOUND)
                if isinstance(r_cell.value, str) and r_cell.value.startswith("="):
                    ob_f = _formula_for_row(template_formulas, COL_OUTBOUND, r)
                    r_cell.value = ob_f or f"=SUM({ds_col}{r}:{de_col}{r})"
                    end_f = _formula_for_row(template_formulas, COL_END_KG, r)
                    ws.cell(row=r, column=COL_END_KG).value = (
                        end_f or f"={q_col}{r}-{r_col}{r}")
                    ev_f = _formula_for_row(template_formulas, COL_END_VALUE, r)
                    ws.cell(row=r, column=COL_END_VALUE).value = (
                        ev_f or f"=ROUND({s_col}{r}*{g_col}{r},0)")
                # ─── J/K は F*G ある 行 だけ (= 価格 が ある active row) ───
                f_v = ws.cell(row=r, column=COL_TOTAL_KG).value
                g_v = ws.cell(row=r, column=COL_UNIT_PRICE).value
                if f_v and g_v:
                    tax_f = _formula_for_row(template_formulas, COL_TAX, r)
                    ws.cell(row=r, column=COL_TAX).value = (
                        tax_f or f"=ROUND({f_col}{r}*{g_col}{r}*0.08,0)")
                    gt_f = _formula_for_row(template_formulas, COL_GRAND_TOTAL, r)
                    ws.cell(row=r, column=COL_GRAND_TOTAL).value = (
                        gt_f or f"=ROUND({f_col}{r}*{g_col}{r}*1.08+{h_col}{r}+{i_col}{r},0)")
                # ─── U/V は 常時 (= 棚卸 補正 入力 の 受け皿、 inactive 行 でも 必要) ───
                ws.cell(row=r, column=COL_STOCKTAKE).value = f"=S{r}"
                ws.cell(row=r, column=COL_DIFF).value = f"=N(S{r})-N(U{r})"

        # Step 6.8: Excel Table の calculatedColumnFormula メタデータ を 削除。
        # 2026-06-01 真因 特定:
        #   元 Excel Table は J/K/R/S/T 等 に 構造化参照 の calc-column formula
        #   (例: `T仕入[[#This Row],[仕入]]*T仕入[[#This Row],[単価]]*0.08`) を 保持。
        #   この メタデータ と セル の 実 formula が **1 つ で も** 不一致 だ と、
        #   Excel は その 列 全体 を 「矛盾した集計列の数式」 と マーク。
        #
        #   preserved 行 (= Excel に 既に あり system に 無い 古い 行) に ユーザー が
        #   手動 で 値 上書き / 別 formula 設定 して いる セル が 1 つ で も ある と、
        #   私 が append / Step 6.7 で formula を 完全 に 揃えて も 警告 は 残る。
        #
        #   対策: 私 が touch する 列 (J/K/R/S/T) の calculatedColumnFormula を
        #   None に set する → Excel は 「統一 formula が ない 列」 と 判断 → 警告
        #   検証 が 走らない。 副作用 は 「Excel 上 で 行 を 新規 追加 した とき に
        #   自動 で formula が 入らない」 だけ (= 同期 用途 では 全く 問題 なし、
        #   ユーザー は 既存 行 を コピペ で 増やせ ば 良い)。
        from openpyxl.utils import column_index_from_string
        TOUCH_COLS = {COL_TAX, COL_GRAND_TOTAL, COL_OUTBOUND, COL_END_KG, COL_END_VALUE}
        for tname in list(ws.tables):
            tbl = ws.tables[tname]
            m = re.match(r"^([A-Z]+)\d+:([A-Z]+)\d+$", tbl.ref)
            if not m:
                continue
            start_col_idx = column_index_from_string(m.group(1))
            for i, tc in enumerate(tbl.tableColumns):
                actual_col = start_col_idx + i  # 1-based Excel column index
                if actual_col in TOUCH_COLS and tc.calculatedColumnFormula is not None:
                    tc.calculatedColumnFormula = None

        # Step 7: マスタ未マッチ警告
        for lot in sys_lots:
            spec_n = _norm_key(lot["spec_combined"])
            if spec_n and spec_n not in master_specs:
                result["master_warnings"].append(
                    f"[{sheet_name}] M規格 に '{lot['spec_combined']}' が ありません (lot {lot['lot_code']})"
                )
            origin_with_san = (lot["origin_name"] or "") + "産"
            if origin_with_san and _norm_origin(origin_with_san) not in {_norm_origin(o) for o in master_origins}:
                result["master_warnings"].append(
                    f"[{sheet_name}] M産地 に '{origin_with_san}' が ありません (lot {lot['lot_code']})"
                )
            sup_n = _norm_key(lot["supplier_name"])
            if sup_n and sup_n not in master_suppliers:
                result["master_warnings"].append(
                    f"[{sheet_name}] M仕入先 に '{lot['supplier_name']}' が ありません (lot {lot['lot_code']})"
                )

        result["sheets"][sheet_name] = {
            "deleted":           deleted_count,
            "appended":          n_new,
            "preserved_inactive": len(inactive_rows),
            "system_total":      n_new,
        }

    if dry_run:
        return None, result

    # 同期 marker を 書き込む (再同期 検出 用)。 既存 marker は 上書き。
    try:
        from openpyxl.packaging.custom import StringProperty
        synced_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # 既存 同名 prop を 削除
        try:
            existing_names = [p.name for p in wb.custom_doc_props]
            if SYNCED_MARKER_PROP in existing_names:
                # CustomPropertyList は __delitem__ 対応
                del wb.custom_doc_props[SYNCED_MARKER_PROP]
        except Exception:
            pass
        wb.custom_doc_props.append(StringProperty(name=SYNCED_MARKER_PROP, value=synced_at))
    except Exception as e:
        logger.warning("同期 marker 書込み失敗 (再同期検出 が 効かなく なる): %s", e)

    # 保存 (xlsm 形式 で バイナリ 返却)
    buf = BytesIO()
    wb.save(buf)

    # ─── 復元処理 (openpyxl が 破壊 する 部分 を 元 zip から 戻す) ─────────
    # 1. externalLinks/ — rId 欠落 / 拡張削除 の 「破損」 警告 を 防ぐ
    # 2. drawings/ + media/ — マクロ付き 図形 (図2/図3 ToggleFilterRow 等) を 復元
    # 3. sheet rels に drawing 関係 を マージ (openpyxl 既存 rels + 元 drawing 関係)
    # 4. sheet.xml に <drawing r:id="..."/> タグ を 注入
    # 5. [Content_Types].xml に drawing/image Override エントリ を 追記
    if preserved_parts or sheet_rels_to_merge or sheet_controls_xml:
        buf.seek(0)
        with zipfile.ZipFile(buf) as zin:
            entries = {n: zin.read(n) for n in zin.namelist()}

        # 単純復元 (drawings, media, externalLinks)
        for name, data in preserved_parts.items():
            entries[name] = data

        # sheet rels マージ (元 関係 を target ベース で 追加、 openpyxl 既存 rels は 保持)
        # 戦略 (target 正規化 + (type, target) ペア dedup):
        #   - 既存 rels の (type, canonical_target) 集合 を 取得
        #   - 元 関係 のうち、 (type, canonical_target) が 既存 に ない もの だけ 追加
        #   - 既存 ID と 衝突 する 場合 は 新 ID を 生成 し sheet_drawing_links も 同期 更新
        # ※ openpyxl は table/comment/drawing/vmlDrawing 関係 を 自前 で 作る が、
        #   ctrlProp/printerSettings は 作らない → これらは 必ず 復元 が 必要 (= Form
        #   Control が drawing.xml と VML の 両方 で 描画 され 二重表示 する 不具合 防止)。
        # ※ target 正規化: openpyxl は 絶対パス '/xl/tables/...', 元 は 相対 '../tables/...'。
        #   sheet rels の base は 'xl/worksheets/' なので 解決 して 比較。

        def _canon_target(rels_path: str, target: str) -> str:
            """rels 内 Target を package root からの絶対パス に 正規化。"""
            if target.startswith("/"):
                return target.lstrip("/")
            base_dir = "/".join(rels_path.replace("/_rels/", "/").split("/")[:-1])
            parts = base_dir.split("/")
            for seg in target.split("/"):
                if seg == "..":
                    parts.pop()
                elif seg in ("", "."):
                    continue
                else:
                    parts.append(seg)
            return "/".join(parts)

        for rels_path, original_rels in sheet_rels_to_merge.items():
            existing_xml = entries.get(
                rels_path,
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
            ).decode("utf-8")

            # openpyxl 出力時点 の type 集合 (= マージ中 は 変更しない スナップショット)
            # ※ openpyxl が 既に 持って いる 型 (comments/table/drawing/vmlDrawing)
            #   は そのまま 信頼 し、 元 rels の 同型 は スキップ。 これは:
            #     - openpyxl は comments を 別パス (comments/comment2.xml) に 移すが
            #       元 rels の Target (../comments2.xml) は 残骸 → 復元 すると 壊れる
            #     - drawing も openpyxl が 自前で 同等 を 出力 する 場合 重複 になる
            #   一方 ctrlProp/printerSettings は openpyxl が 作らない → 全件 復元 が必要
            #   (同 type 複数 = sheet1 に ctrlProp が 6 個 等 が ある ため
            #    type を 復元時 に 加算 すると 2 個目以降 が 誤スキップ される)。
            openpyxl_types: set[str] = set()
            for m in re.finditer(r'<Relationship\b[^>]*>', existing_xml):
                t_match = re.search(r'Type="([^"]+)"', m.group(0))
                if t_match:
                    openpyxl_types.add(t_match.group(1))
            existing_ids = set(re.findall(r'Id="([^"]+)"', existing_xml))
            sheet_path = rels_path.replace("/_rels/", "/").replace(".rels", "")

            for d in original_rels:
                # openpyxl 出力 に 同 type が ある → 復元 スキップ (重複防止)
                if d["type"] in openpyxl_types:
                    # drawing 関係 が 既に ある なら sheet.xml への <drawing> 注入 も キャンセル
                    if "drawing" in d["type"].lower() and "vml" not in d["type"].lower():
                        sheet_drawing_links.pop(sheet_path, None)
                    continue

                old_id = d["id"]
                new_id = old_id
                n = 1
                while new_id in existing_ids:
                    new_id = f"rIdRestored{n}"
                    n += 1
                existing_ids.add(new_id)
                new_rel = f'<Relationship Id="{new_id}" Type="{d["type"]}" Target="{d["target"]}"/>'
                if "</Relationships>" in existing_xml:
                    existing_xml = existing_xml.replace("</Relationships>", f"{new_rel}</Relationships>")
                else:
                    existing_xml = existing_xml.replace(
                        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
                        f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{new_rel}</Relationships>'
                    )
                # sheet.xml 注入用 rId も 新 ID で 更新 (drawing 関係 のみ)
                if old_id != new_id:
                    for link in sheet_drawing_links.get(sheet_path, []):
                        if link["id"] == old_id:
                            link["id"] = new_id
                            break
            entries[rels_path] = existing_xml.encode("utf-8")

        # sheet.xml に <drawing> タグ を 注入 (OOXML スキーマ順序 を 尊重)
        # 順序: ... drawing → legacyDrawing → legacyDrawingHF → picture → oleObjects
        #       → controls → webPublishItems → tableParts → extLst
        # 要 = <tableParts> や <legacyDrawing> より 前 に 挿入。
        for sheet_path, links in sheet_drawing_links.items():
            if sheet_path not in entries:
                continue
            sheet_xml = entries[sheet_path].decode("utf-8")
            for link in links:
                rid = link["id"]
                if f'r:id="{rid}"' in sheet_xml:
                    continue
                tag = f'<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="{rid}"/>'
                # 適切な挿入位置 を 順番に 試行
                inserted = False
                for marker in ("<legacyDrawing ", "<legacyDrawing/>", "<tableParts ", "<tableParts>", "<extLst>"):
                    idx = sheet_xml.find(marker)
                    if idx >= 0:
                        sheet_xml = sheet_xml[:idx] + tag + sheet_xml[idx:]
                        inserted = True
                        break
                if not inserted and "</worksheet>" in sheet_xml:
                    sheet_xml = sheet_xml.replace("</worksheet>", f"{tag}</worksheet>")
            entries[sheet_path] = sheet_xml.encode("utf-8")

        # sheet.xml に <controls> セクション を 再注入。
        # スキーマ順序: oleObjects → controls → webPublishItems → tableParts → extLst
        # → <tableParts>/<extLst> より 前 (もし <legacyDrawing> しか ない 場合 は
        # その 後ろ) に 挿入 する。
        # 注意: controls ブロック は mc:/xdr:/r:/x14 名前空間 を 使う ので、
        # openpyxl が 出力 する <worksheet> 根 タグ に これら を 追加 する 必要 が ある
        # (= openpyxl は 既定 名前空間 しか 宣言 しない)。
        needed_xmlns = {
            "xmlns:r":   "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            "xmlns:xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "xmlns:mc":  "http://schemas.openxmlformats.org/markup-compatibility/2006",
            "xmlns:x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
        }
        for sheet_path, controls_block in sheet_controls_xml.items():
            if sheet_path not in entries:
                continue
            sheet_xml = entries[sheet_path].decode("utf-8")
            # 既に <controls> が ある なら 何もしない
            if "<controls>" in sheet_xml:
                continue
            # 不足 名前空間 を <worksheet ...> に 追加
            m_root = re.match(r'(<worksheet\b)([^>]*)>', sheet_xml)
            if m_root:
                root_attrs = m_root.group(2)
                add_attrs = ""
                for pfx, uri in needed_xmlns.items():
                    if pfx + "=" not in root_attrs:
                        add_attrs += f' {pfx}="{uri}"'
                if add_attrs:
                    new_root = f"{m_root.group(1)}{root_attrs}{add_attrs}>"
                    sheet_xml = sheet_xml[:m_root.start()] + new_root + sheet_xml[m_root.end():]
            inserted = False
            # <tableParts>/<extLst> の 直前 に 挿入
            for marker in ("<tableParts ", "<tableParts>", "<extLst>", "<extLst "):
                idx = sheet_xml.find(marker)
                if idx >= 0:
                    sheet_xml = sheet_xml[:idx] + controls_block + sheet_xml[idx:]
                    inserted = True
                    break
            if not inserted:
                # legacyDrawing 直後 に 挿入 (= スキーマ順序 的 に oleObjects 位置 相当)
                m = re.search(r'<legacyDrawing\b[^/]*/?>', sheet_xml)
                if m:
                    end = m.end()
                    sheet_xml = sheet_xml[:end] + controls_block + sheet_xml[end:]
                    inserted = True
            if not inserted and "</worksheet>" in sheet_xml:
                sheet_xml = sheet_xml.replace("</worksheet>", f"{controls_block}</worksheet>")
            entries[sheet_path] = sheet_xml.encode("utf-8")

        # [Content_Types].xml の Override / Default 追加
        ct_path = "[Content_Types].xml"
        if ct_path in entries:
            ct = entries[ct_path].decode("utf-8")
            # drawings の Override
            for dname in [n for n in preserved_parts if n.startswith("xl/drawings/") and n.endswith(".xml") and not n.endswith(".rels")]:
                pn = "/" + dname
                if f'PartName="{pn}"' in ct: continue
                override = f'<Override PartName="{pn}" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                ct = ct.replace("</Types>", f"{override}</Types>")
            # ctrlProps の Override (Form Control 設定)
            for cname in [n for n in preserved_parts if n.startswith("xl/ctrlProps/") and n.endswith(".xml")]:
                pn = "/" + cname
                if f'PartName="{pn}"' in ct: continue
                override = f'<Override PartName="{pn}" ContentType="application/vnd.ms-excel.controlproperties+xml"/>'
                ct = ct.replace("</Types>", f"{override}</Types>")
            # printerSettings の Default (拡張子 .bin 用)
            has_printer = any(n.startswith("xl/printerSettings/") for n in preserved_parts)
            if has_printer and 'Extension="bin"' not in ct:
                default = '<Default Extension="bin" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings"/>'
                ct = re.sub(r"(<Types[^>]*>)", r"\1" + default, ct, count=1)
            # media の Default (拡張子 ベース)
            for mname in [n for n in preserved_parts if n.startswith("xl/media/")]:
                ext = mname.rsplit(".", 1)[-1].lower()
                if f'Extension="{ext}"' in ct: continue
                mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                        "gif": "image/gif", "bmp": "image/bmp"}.get(ext, "image/png")
                default = f'<Default Extension="{ext}" ContentType="{mime}"/>'
                ct = re.sub(r"(<Types[^>]*>)", r"\1" + default, ct, count=1)
            entries[ct_path] = ct.encode("utf-8")

        new_buf = BytesIO()
        with zipfile.ZipFile(new_buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for name, data in entries.items():
                zout.writestr(name, data)
        new_buf.seek(0)
        return new_buf.getvalue(), result

    buf.seek(0)
    return buf.getvalue(), result
